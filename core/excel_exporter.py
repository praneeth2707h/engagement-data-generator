"""Stage 9 — Excel Workbook Export.

ExcelExporter converts simulation outputs from all upstream stages into a
business-ready Excel workbook (.xlsx).  The workbook contains six sheets:

  Sheet 1 — Event Data            Normalised per-event log.
  Sheet 2 — Campaign Metrics      Aggregated KPIs from the metrics DataFrame.
  Sheet 3 — Validation Results    Per-rule validation outcomes.
  Sheet 4 — Validation Summary    Category-level validation aggregates.
  Sheet 5 — Realism Report        Requested vs. actual rate comparison.
  Sheet 6 — Diagnostics           Requested/actual/variance diagnostic table.

Architecture references
-----------------------
* ARCH-003  — Stage 9 in the 11-stage pipeline
* ARCH-011  — No iterrows(); all processing vectorised

Formatting conventions
----------------------
* All sheets: freeze top row, auto-fit column widths, Table-styled (TableStyleMedium9).
* Numeric cells: consistent decimal places (2dp for rates, 0dp for counts).
* Output is deterministic — sheet and column order are fixed.
"""
from __future__ import annotations

import io
import re
import zipfile
from datetime import date, datetime
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from models.config_registry import ConfigRegistry
from utils.exceptions import InputValidationError
from utils.logger import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Sheet names (fixed order — do not reorder)
# ---------------------------------------------------------------------------
_SHEET_EVENT_DATA    = "Event Data"
_SHEET_CAMPAIGN      = "Campaign Metrics"
_SHEET_VAL_RESULTS   = "Validation Results"
_SHEET_VAL_SUMMARY   = "Validation Summary"
_SHEET_REALISM       = "Realism Report"
_SHEET_DIAGNOSTICS   = "Diagnostics"

SHEET_NAMES: tuple[str, ...] = (
    _SHEET_EVENT_DATA,
    _SHEET_CAMPAIGN,
    _SHEET_VAL_RESULTS,
    _SHEET_VAL_SUMMARY,
    _SHEET_REALISM,
    _SHEET_DIAGNOSTICS,
)

# ---------------------------------------------------------------------------
# Sheet 1 column mapping  (events_df col → display name)
# ---------------------------------------------------------------------------
_EVENT_COL_MAP: dict[str, str] = {
    "simulation_date": "Date",
    "user_id":         "User_ID",
    "action_type":     "Action",
    "channel":         "Channel",
    "vendor":          "Vendor",
    "current_ad":      "Creative",
    "journey_stage":   "Journey_Stage",
    "segment":         "Segment",
    "trigger_name":    "Trigger_Name",
}
_EVENT_DISPLAY_COLS: tuple[str, ...] = (
    "Date", "User_ID", "Action", "Channel",
    "Vendor", "Creative", "Journey_Stage", "Segment", "Trigger_Name",
)

# ---------------------------------------------------------------------------
# Sheet 2 column names
# ---------------------------------------------------------------------------
_METRICS_DISPLAY_COLS: tuple[str, ...] = (
    "Metric", "Value",
)

# ---------------------------------------------------------------------------
# Sheet 6 column names
# ---------------------------------------------------------------------------
_DIAG_DISPLAY_COLS: tuple[str, ...] = (
    "Metric", "Entity", "Requested", "Actual", "Variance", "Variance_%",
)

# ---------------------------------------------------------------------------
# Table style applied to every sheet
# ---------------------------------------------------------------------------
_TABLE_STYLE = "TableStyleMedium9"

# ---------------------------------------------------------------------------
# Header fill colour (hex, no #)
# ---------------------------------------------------------------------------
_HEADER_FILL_HEX = "1F4E79"  # dark navy
_HEADER_FONT_HEX = "FFFFFF"  # white


class ExcelExporter:
    """Stage 9 — Simulation-output Excel workbook exporter.

    Accepts the six DataFrames produced by the simulation pipeline and emits
    a formatted, table-styled Excel workbook.

    Args:
        config: Campaign ConfigRegistry.  Used for campaign metadata (name,
            dates) written into the Campaign Metrics sheet header row.
    """

    def __init__(self, config: ConfigRegistry) -> None:
        """Initialise the exporter with campaign configuration.

        Args:
            config: ConfigRegistry for the current campaign run.
        """
        self._config = config
        logger.info(
            "ExcelExporter initialised — campaign: %s", config.campaign_id
        )

    # -----------------------------------------------------------------------
    # Public API
    # -----------------------------------------------------------------------

    def export(
        self,
        events_df: pd.DataFrame,
        state_df: pd.DataFrame,
        metrics_df: pd.DataFrame,
        validation_results_df: pd.DataFrame,
        validation_summary_df: pd.DataFrame,
        realism_report_df: pd.DataFrame,
        output_path: str | None = None,
    ) -> bytes:
        """Build and optionally save the Excel workbook.

        Assembles all six sheets, applies formatting, and either writes the
        workbook to *output_path* or returns it as a ``bytes`` object (useful
        for Streamlit ``st.download_button``).

        Args:
            events_df: Events DataFrame from EngagementGenerator.
            state_df: State DataFrame from UserStateManager / AudienceManager.
            metrics_df: Metrics DataFrame from EngagementGenerator.
            validation_results_df: Per-rule results from ValidationEngine.
            validation_summary_df: Category summary from ValidationEngine.
            realism_report_df: Realism report from ValidationEngine.
            output_path: If provided, the workbook is also saved to this path.
                If None, only the bytes representation is returned.

        Returns:
            Raw bytes of the .xlsx workbook.

        Raises:
            InputValidationError: If events_df is missing required columns.
        """
        _validate_events_columns(events_df)

        logger.info(
            "ExcelExporter.export() — %d events, %d state rows, "
            "%d metric rows, %d validation rules",
            len(events_df), len(state_df), len(metrics_df),
            len(validation_results_df),
        )

        wb = Workbook()
        # Remove the default empty sheet
        wb.remove(wb.active)
        # Pin workbook timestamps to a fixed epoch so output is byte-for-byte
        # deterministic across runs (openpyxl otherwise embeds wall-clock time
        # in docProps/core.xml, breaking @st.cache_data and test equality).
        from datetime import datetime as _dt
        _EPOCH = _dt(2000, 1, 1, 0, 0, 0)
        wb.properties.created  = _EPOCH
        wb.properties.modified = _EPOCH

        self._write_event_data(wb, events_df, state_df)
        self._write_campaign_metrics(wb, metrics_df, events_df, state_df)
        self._write_dataframe_sheet(wb, _SHEET_VAL_RESULTS,  validation_results_df)
        self._write_dataframe_sheet(wb, _SHEET_VAL_SUMMARY,  validation_summary_df)
        self._write_dataframe_sheet(wb, _SHEET_REALISM,       realism_report_df)
        self._write_diagnostics(wb, events_df, state_df, metrics_df)

        buf = io.BytesIO()
        wb.save(buf)
        raw = _normalize_workbook_bytes(buf.getvalue())

        if output_path:
            with open(output_path, "wb") as fh:
                fh.write(raw)
            logger.info("ExcelExporter: workbook saved to %s (%d bytes)", output_path, len(raw))

        logger.info("ExcelExporter.export() complete — %d bytes", len(raw))
        return raw

    # -----------------------------------------------------------------------
    # Sheet writers
    # -----------------------------------------------------------------------

    def _write_event_data(
        self,
        wb: Workbook,
        events_df: pd.DataFrame,
        state_df: pd.DataFrame,
    ) -> None:
        """Write Sheet 1 — Event Data.

        Normalises column names, fills missing optional columns (journey_stage,
        segment, trigger_name, vendor) with empty strings, then writes.
        """
        ws = wb.create_sheet(_SHEET_EVENT_DATA)

        df = events_df.copy()

        # Merge journey_stage from state_df if not in events and state has it
        if "journey_stage" not in df.columns:
            if "current_ad" in state_df.columns and "user_id" in state_df.columns:
                stage_map = (
                    state_df[["user_id", "current_ad"]]
                    .drop_duplicates("user_id")
                    .rename(columns={"current_ad": "journey_stage"})
                )
                if "user_id" in df.columns:
                    df = df.merge(stage_map, on="user_id", how="left")
            if "journey_stage" not in df.columns:
                df["journey_stage"] = ""

        # Ensure all optional columns exist
        for src_col in _EVENT_COL_MAP:
            if src_col not in df.columns:
                df[src_col] = ""

        # Convert Categorical → str
        for col in ("trigger_name", "segment", "vendor"):
            if col in df.columns and hasattr(df[col], "cat"):
                df[col] = df[col].astype(str)

        # Select and rename
        df = df[[c for c in _EVENT_COL_MAP if c in df.columns]].rename(
            columns=_EVENT_COL_MAP
        )

        # Ensure all display columns are present (fill missing)
        for dcol in _EVENT_DISPLAY_COLS:
            if dcol not in df.columns:
                df[dcol] = ""

        df = df[list(_EVENT_DISPLAY_COLS)]

        # Format date column
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")

        _write_df_to_sheet(ws, df, table_name="tblEventData")

    def _write_campaign_metrics(
        self,
        wb: Workbook,
        metrics_df: pd.DataFrame,
        events_df: pd.DataFrame,
        state_df: pd.DataFrame,
    ) -> None:
        """Write Sheet 2 — Campaign Metrics.

        Aggregates totals/rates from metrics_df and computes supplementary
        metrics (TER, segment engagement rate, average frequency) from
        events_df and state_df.
        """
        ws = wb.create_sheet(_SHEET_CAMPAIGN)

        rows = _compute_campaign_metrics(
            metrics_df, events_df, state_df, self._config
        )
        df = pd.DataFrame(rows, columns=["Metric", "Value"])
        _write_df_to_sheet(ws, df, table_name="tblCampaignMetrics")

    def _write_dataframe_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        df: pd.DataFrame,
    ) -> None:
        """Write an arbitrary DataFrame to a new sheet with table formatting."""
        ws   = wb.create_sheet(sheet_name)
        safe = _safe_table_name(sheet_name)
        _write_df_to_sheet(ws, df, table_name=safe)

    def _write_diagnostics(
        self,
        wb: Workbook,
        events_df: pd.DataFrame,
        state_df: pd.DataFrame,
        metrics_df: pd.DataFrame,
    ) -> None:
        """Write Sheet 6 — Diagnostics.

        Produces a Requested/Actual/Variance table covering CTR per ad,
        TER per trigger, and segment mix per segment.
        """
        ws   = wb.create_sheet(_SHEET_DIAGNOSTICS)
        rows = _compute_diagnostics(events_df, state_df, metrics_df, self._config)
        df   = pd.DataFrame(rows, columns=list(_DIAG_DISPLAY_COLS))
        _write_df_to_sheet(ws, df, table_name="tblDiagnostics")


# ---------------------------------------------------------------------------
# Metric computation helpers
# ---------------------------------------------------------------------------

def _compute_campaign_metrics(
    metrics_df: pd.DataFrame,
    events_df: pd.DataFrame,
    state_df: pd.DataFrame,
    config: ConfigRegistry,
) -> list[tuple[str, Any]]:
    """Compute the nine required Campaign Metrics rows.

    Args:
        metrics_df: Daily metrics from EngagementGenerator.
        events_df:  Events DataFrame.
        state_df:   State DataFrame.
        config:     ConfigRegistry for TER and segment targets.

    Returns:
        List of (Metric, Value) tuples in display order.
    """
    # ── Totals from metrics_df ─────────────────────────────────────────────
    if not metrics_df.empty:
        total_impressions = int(metrics_df["n_impressions"].sum()) if "n_impressions" in metrics_df.columns else 0
        total_clicks      = int(metrics_df["n_clicks"].sum())      if "n_clicks"      in metrics_df.columns else 0
        total_sends       = int(metrics_df["n_sends"].sum())       if "n_sends"       in metrics_df.columns else 0
        total_opens       = int(metrics_df["n_opens"].sum())       if "n_opens"       in metrics_df.columns else 0
    else:
        total_impressions = total_clicks = total_sends = total_opens = 0

    # ── CTR ───────────────────────────────────────────────────────────────
    ctr = round(total_clicks / total_impressions, 4) if total_impressions > 0 else 0.0

    # ── Open rate ─────────────────────────────────────────────────────────
    open_rate = round(total_opens / total_sends, 4) if total_sends > 0 else 0.0

    # ── Trigger Engagement Rate ────────────────────────────────────────────
    ter = _compute_ter(events_df, state_df)

    # ── Segment Engagement Rate ────────────────────────────────────────────
    ser = _compute_ser(events_df, state_df)

    # ── Average Frequency ─────────────────────────────────────────────────
    avg_freq = _compute_avg_frequency(events_df)

    return [
        ("Total Impressions",         total_impressions),
        ("Total Clicks",              total_clicks),
        ("CTR",                       f"{ctr:.4f}"),
        ("Total Sends",               total_sends),
        ("Total Opens",               total_opens),
        ("Open Rate",                 f"{open_rate:.4f}"),
        ("Trigger Engagement Rate",   f"{ter:.4f}"),
        ("Segment Engagement Rate",   f"{ser:.4f}"),
        ("Average Frequency",         f"{avg_freq:.2f}"),
    ]


def _compute_ter(events_df: pd.DataFrame, state_df: pd.DataFrame) -> float:
    """Overall Trigger Engagement Rate = qualifying events / triggered users."""
    if events_df.empty:
        return 0.0
    n_triggered = len(state_df) if state_df is not None and not state_df.empty else 0
    if n_triggered == 0:
        return 0.0
    qual = _qualifying_mask(events_df)
    n_engaged = int(events_df.loc[qual, "user_id"].nunique()) if qual.any() else 0
    return round(n_engaged / n_triggered, 4)


def _compute_ser(events_df: pd.DataFrame, state_df: pd.DataFrame) -> float:
    """Segment Engagement Rate = engaged users with known segment / total engaged."""
    if events_df.empty:
        return 0.0
    qual = _qualifying_mask(events_df)
    if not qual.any():
        return 0.0
    q_df = events_df.loc[qual]
    if "segment" not in q_df.columns:
        return 0.0
    seg_vals = q_df["segment"].astype(str)
    has_seg  = seg_vals.str.strip().ne("").values
    n_total  = int(q_df["user_id"].nunique())
    if n_total == 0:
        return 0.0
    n_with_seg = int(q_df.loc[has_seg, "user_id"].nunique())
    return round(n_with_seg / n_total, 4)


def _compute_avg_frequency(events_df: pd.DataFrame) -> float:
    """Average number of qualifying events per engaged user."""
    if events_df.empty:
        return 0.0
    qual = _qualifying_mask(events_df)
    if not qual.any():
        return 0.0
    q_df   = events_df.loc[qual]
    counts = q_df.groupby("user_id").size()
    return round(float(counts.mean()), 4) if not counts.empty else 0.0


def _compute_diagnostics(
    events_df: pd.DataFrame,
    state_df: pd.DataFrame,
    metrics_df: pd.DataFrame,
    config: ConfigRegistry,
) -> list[tuple[str, str, float, float, float, float]]:
    """Build the Diagnostics sheet rows.

    Returns list of (Metric, Entity, Requested, Actual, Variance, Variance_%).
    Covers: CTR per ad, TER per trigger, Segment Mix per segment.
    """
    rows: list[tuple[str, str, float, float, float, float]] = []
    eps = 1e-10

    # ── CTR per ad ────────────────────────────────────────────────────────
    if not events_df.empty and "current_ad" in events_df.columns:
        imp_mask   = events_df["action_type"].astype(str) == "Impression"
        click_mask = events_df["action_type"].astype(str) == "Click"
        imp_counts   = events_df.loc[imp_mask,   "current_ad"].astype(str).value_counts()
        click_counts = events_df.loc[click_mask,  "current_ad"].astype(str).value_counts()

        for ad in config.ads:
            requested = float(ad.target_ctr) if ad.target_ctr is not None else 0.0
            n_imp     = int(imp_counts.get(ad.ad_name, 0))
            n_click   = int(click_counts.get(ad.ad_name, 0))
            actual    = round(n_click / n_imp, 4) if n_imp > 0 else 0.0
            variance  = round(actual - requested, 4)
            var_pct   = round(variance / max(abs(requested), eps) * 100, 2)
            rows.append(("CTR", ad.ad_name, requested, actual, variance, var_pct))
    else:
        for ad in config.ads:
            requested = float(ad.target_ctr) if ad.target_ctr is not None else 0.0
            rows.append(("CTR", ad.ad_name, requested, 0.0, -requested,
                         round(-100.0 if abs(requested) > eps else 0.0, 2)))

    # ── TER per trigger ───────────────────────────────────────────────────
    if not events_df.empty and "trigger_name" in events_df.columns:
        qual     = _qualifying_mask(events_df)
        q_df     = events_df.loc[qual] if qual.any() else events_df.iloc[0:0]
        trig_col = q_df["trigger_name"].astype(str) if not q_df.empty else pd.Series(dtype=str)

        # Users per trigger from state_df
        if state_df is not None and not state_df.empty and "trigger_name" in state_df.columns:
            trig_users = (
                state_df["trigger_name"]
                .astype(str)
                .value_counts()
            )
        else:
            trig_users = pd.Series(dtype=int)

        for trig in config.triggers:
            requested  = float(trig.engagement_rate_target)
            n_users    = int(trig_users.get(trig.trigger_name, 0))
            n_engaged  = int(trig_col[trig_col == trig.trigger_name].count()) if not trig_col.empty else 0
            actual     = round(n_engaged / n_users, 4) if n_users > 0 else 0.0
            variance   = round(actual - requested, 4)
            var_pct    = round(variance / max(abs(requested), eps) * 100, 2)
            rows.append(("TER", trig.trigger_name, requested, actual, variance, var_pct))
    else:
        for trig in config.triggers:
            requested = float(trig.engagement_rate_target)
            rows.append(("TER", trig.trigger_name, requested, 0.0, -requested,
                         round(-100.0 if abs(requested) > eps else 0.0, 2)))

    # ── Segment mix ───────────────────────────────────────────────────────
    if config.segments:
        if state_df is not None and not state_df.empty and "segment" in state_df.columns:
            seg_counts = state_df["segment"].astype(str).value_counts()
            n_total    = len(state_df)
        else:
            seg_counts = pd.Series(dtype=int)
            n_total    = 0

        for seg in config.segments:
            requested  = round(seg.distribution_pct / 100.0, 4)
            n_seg      = int(seg_counts.get(seg.segment_name, 0))
            actual     = round(n_seg / n_total, 4) if n_total > 0 else 0.0
            variance   = round(actual - requested, 4)
            var_pct    = round(variance / max(abs(requested), eps) * 100, 2)
            rows.append(("Segment Mix", seg.segment_name, requested, actual, variance, var_pct))

    return rows


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------

def _write_df_to_sheet(
    ws: Any,
    df: pd.DataFrame,
    table_name: str,
) -> None:
    """Write a DataFrame to an openpyxl worksheet with full formatting.

    Steps:
      1. Write header row with styled fill/font.
      2. Write data rows.
      3. Auto-fit column widths.
      4. Freeze top row.
      5. Add openpyxl Table object for Excel table formatting.

    Args:
        ws: openpyxl Worksheet.
        df: DataFrame to write.
        table_name: Internal Excel table name (must be unique across workbook,
            letters/underscores only).
    """
    if df is None:
        df = pd.DataFrame()

    # Sanitise: replace NaN/None with empty string for clean display
    df = df.fillna("").copy()

    # Convert any remaining non-string objects that openpyxl can't handle
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].astype(str).replace("nan", "").replace("<NA>", "")

    n_rows, n_cols = df.shape

    # ── Header row (row 1) ─────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL_HEX)
    header_font = Font(bold=True, color=_HEADER_FONT_HEX)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # ── Column widths (auto-fit to content) ───────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        # Maximum of header length and max data length, capped at 60
        header_len = len(str(col_name))
        if n_rows > 0:
            col_series = df.iloc[:, col_idx - 1].astype(str)
            max_data   = col_series.str.len().max() if not col_series.empty else 0
        else:
            max_data = 0
        width = min(max(header_len, int(max_data), 8) + 2, 60)
        ws.column_dimensions[col_letter].width = width

    # ── Freeze top row ─────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Excel Table ────────────────────────────────────────────────────────
    if n_cols > 0:
        last_col  = get_column_letter(n_cols)
        last_row  = max(n_rows + 1, 2)   # table must include header even if empty data
        ref       = f"A1:{last_col}{last_row}"
        tbl       = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name=_TABLE_STYLE,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)


def _qualifying_mask(events_df: pd.DataFrame) -> pd.Series:
    """Return a boolean mask identifying qualifying engagement actions.

    Qualifying = Display→Click, Email→{Open,Click}, WhatsApp→{Open,Click}.
    Mirrors the same function in ValidationEngine / EngagementGenerator.
    """
    if events_df.empty or "action_type" not in events_df.columns:
        return pd.Series(False, index=events_df.index)

    ch  = events_df["channel"].astype(str) if "channel" in events_df.columns else pd.Series("", index=events_df.index)
    act = events_df["action_type"].astype(str)

    _DISPLAY = frozenset({"Display", "Endemic_Display", "Programmatic_Display", "Banner"})
    disp  = ch.isin(_DISPLAY) & (act == "Click")
    email = (ch == "Email")     & act.isin(["Open", "Click"])
    wa    = (ch == "WhatsApp")  & act.isin(["Open", "Click"])
    return disp | email | wa


def _safe_table_name(sheet_name: str) -> str:
    """Convert a sheet name to a valid Excel table name (alphanumeric + underscore)."""
    return "tbl" + "".join(c if c.isalnum() else "_" for c in sheet_name)


def _validate_events_columns(events_df: pd.DataFrame) -> None:
    """Raise InputValidationError if required events columns are missing."""
    required = {"user_id", "simulation_date", "channel", "action_type"}
    missing  = required - set(events_df.columns)
    if missing:
        raise InputValidationError(
            "events_df",
            f"Missing required columns: {sorted(missing)}",
        )


__all__ = ["ExcelExporter", "SHEET_NAMES"]


# ---------------------------------------------------------------------------
# Workbook determinism helper  (DEF-EX-002)
# ---------------------------------------------------------------------------

# openpyxl/writer/excel.py line 292 forcibly overwrites
# ``workbook.properties.modified = datetime.datetime.now()`` during save(),
# so the _EPOCH assignment before save() is insufficient.  We fix this by
# post-processing the raw ZIP bytes:
#   1. Re-writing every ZIP local-file-header date_time to a fixed epoch.
#   2. Replacing the dcterms:modified value in docProps/core.xml with the same
#      epoch string.
# Both changes ensure byte-for-byte identical output across runs when the
# input DataFrames are identical.

_ZIP_EPOCH     = (2000, 1, 1, 0, 0, 0)
_MODIFIED_EPOCH_XML = (
    b'<dcterms:modified '
    b'xmlns:dcterms="http://purl.org/dc/terms/" '
    b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
    b'xsi:type="dcterms:W3CDTF">2000-01-01T00:00:00Z</dcterms:modified>'
)
_MODIFIED_RE = re.compile(
    rb'<dcterms:modified[^>]*>[^<]*</dcterms:modified>'
)


def _normalize_workbook_bytes(data: bytes) -> bytes:
    """Return a deterministic rewrite of raw .xlsx (ZIP) bytes.

    Applies two normalizations:
    * All ZIP entry ``date_time`` fields → ``_ZIP_EPOCH``.
    * ``dcterms:modified`` in ``docProps/core.xml`` → epoch ISO string.

    The re-compression uses ``zipfile.ZIP_DEFLATED`` at the default zlib
    level, which is deterministic for identical input content.

    Args:
        data: Raw bytes from ``Workbook.save()``.

    Returns:
        Deterministic bytes suitable for byte-level equality comparison.
    """
    buf_in  = io.BytesIO(data)
    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_in, "r") as zf_in:
        with zipfile.ZipFile(
            buf_out, "w",
            compression=zipfile.ZIP_DEFLATED,
            allowZip64=True,
        ) as zf_out:
            for info in zf_in.infolist():
                info.date_time = _ZIP_EPOCH
                content = zf_in.read(info.filename)
                if info.filename == "docProps/core.xml":
                    content = _MODIFIED_RE.sub(_MODIFIED_EPOCH_XML, content)
                zf_out.writestr(info, content)
    return buf_out.getvalue()
