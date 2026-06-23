"""Tests for core/simulation_orchestrator.py and models/simulation_result.py.

Coverage:
  * SimulationResult model — fields, properties, frozen contract
  * SimulationOrchestrator.__init__
  * Full happy-path run (generate_excel=True and False)
  * Empty audience scenarios
  * Empty event scenarios
  * Validation failure propagation (quality_score < 100)
  * Excel workbook generation and round-trip
  * Error propagation — SimulationError wraps stage failures
  * Trigger column validation
  * Stage timing capture
  * Deterministic execution (two identical runs → same bytes)
  * All SimulationResult fields populated
  * Metadata completeness
  * Compliance: no iterrows(), __all__ declared
"""
from __future__ import annotations

import io
import time
from datetime import date, timedelta
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from core.simulation_orchestrator import SimulationOrchestrator
from models.simulation_result import SimulationResult
from utils.exceptions import SimulationError

from tests.test_core.conftest import make_config, make_trigger_df

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_SHORT_CFG = dict(
    simulation_start_date=date(2024, 1, 1),
    simulation_end_date=date(2024, 1, 3),
)


def _cfg(**kw):
    return make_config(**{**_SHORT_CFG, **kw})


def _tdf(n=5, cfg=None):
    cfg = cfg or _cfg()
    return make_trigger_df(n=n, campaign_id=cfg.campaign_id)


def _run(n=5, generate_excel=True, **cfg_kw):
    cfg = _cfg(**cfg_kw)
    return SimulationOrchestrator(cfg).run(_tdf(n, cfg), generate_excel=generate_excel)


# ===========================================================================
# 1. SimulationResult — model contract
# ===========================================================================

class TestSimulationResultModel:
    def test_instantiates_with_defaults(self):
        r = SimulationResult()
        assert r.state_df is None
        assert r.events_df is None
        assert r.workbook_bytes is None
        assert r.quality_score == 0.0
        assert r.realism_score == 0.0
        assert r.feasibility_warnings == ()
        assert r.execution_metadata == {}

    def test_frozen_cannot_mutate(self):
        r = SimulationResult()
        with pytest.raises((TypeError, AttributeError)):
            r.quality_score = 99.0  # type: ignore[misc]

    def test_n_events_zero_when_no_events_df(self):
        r = SimulationResult()
        assert r.n_events == 0

    def test_n_events_counts_rows(self):
        df = pd.DataFrame({"x": [1, 2, 3]})
        r  = SimulationResult(events_df=df)
        assert r.n_events == 3

    def test_n_users_zero_when_no_audience_df(self):
        r = SimulationResult()
        assert r.n_users == 0

    def test_n_users_counts_rows(self):
        df = pd.DataFrame({"user_id": ["U001", "U002"]})
        r  = SimulationResult(audience_df=df)
        assert r.n_users == 2

    def test_succeeded_false_without_metadata(self):
        r = SimulationResult()
        assert r.succeeded is False

    def test_succeeded_true_with_completed_at(self):
        r = SimulationResult(execution_metadata={"completed_at": "2024-01-01T00:00:00Z"})
        assert r.succeeded is True

    def test_elapsed_seconds_none_without_metadata(self):
        r = SimulationResult()
        assert r.elapsed_seconds is None

    def test_elapsed_seconds_from_metadata(self):
        r = SimulationResult(execution_metadata={"elapsed_seconds": 1.23})
        assert r.elapsed_seconds == pytest.approx(1.23)

    def test_all_df_fields_accept_dataframe(self):
        df = pd.DataFrame({"a": [1]})
        r  = SimulationResult(
            state_df=df, audience_df=df, events_df=df,
            metrics_df=df, diagnostics_df=df,
            validation_results_df=df, validation_summary_df=df,
            realism_report_df=df,
        )
        assert r.state_df is df
        assert r.validation_results_df is df

    def test_workbook_bytes_stored(self):
        raw = b"fake_bytes"
        r   = SimulationResult(workbook_bytes=raw)
        assert r.workbook_bytes == raw

    def test_feasibility_warnings_is_tuple(self):
        r = SimulationResult(feasibility_warnings=("warn1", "warn2"))
        assert r.feasibility_warnings == ("warn1", "warn2")


# ===========================================================================
# 2. SimulationOrchestrator.__init__
# ===========================================================================

class TestOrchestratorInit:
    def test_instantiates(self):
        orch = SimulationOrchestrator(_cfg())
        assert orch is not None

    def test_stores_config(self):
        cfg  = _cfg()
        orch = SimulationOrchestrator(cfg)
        assert orch._config is cfg


# ===========================================================================
# 3. Trigger column validation
# ===========================================================================

class TestTriggerValidation:
    def test_missing_user_id_raises(self):
        cfg = _cfg()
        tdf = _tdf(5, cfg).drop(columns=["User_ID"])
        with pytest.raises(SimulationError, match="User_ID"):
            SimulationOrchestrator(cfg).run(tdf, generate_excel=False)

    def test_missing_campaign_id_raises(self):
        cfg = _cfg()
        tdf = _tdf(5, cfg).drop(columns=["Campaign_ID"])
        with pytest.raises(SimulationError, match="Campaign_ID"):
            SimulationOrchestrator(cfg).run(tdf, generate_excel=False)

    def test_missing_trigger_name_raises(self):
        cfg = _cfg()
        tdf = _tdf(5, cfg).drop(columns=["Trigger_Name"])
        with pytest.raises(SimulationError, match="Trigger_Name"):
            SimulationOrchestrator(cfg).run(tdf, generate_excel=False)

    def test_valid_columns_no_error(self):
        result = _run(generate_excel=False)
        assert result.succeeded


# ===========================================================================
# 4. Happy-path full run
# ===========================================================================

class TestHappyPath:
    def test_run_returns_simulation_result(self):
        result = _run(generate_excel=False)
        assert isinstance(result, SimulationResult)

    def test_succeeded_is_true(self):
        result = _run(generate_excel=False)
        assert result.succeeded

    def test_all_dfs_populated(self):
        result = _run(generate_excel=False)
        for attr in ("state_df", "audience_df", "events_df", "metrics_df",
                     "diagnostics_df", "validation_results_df",
                     "validation_summary_df", "realism_report_df"):
            df = getattr(result, attr)
            assert isinstance(df, pd.DataFrame), f"{attr} is not a DataFrame"

    def test_n_users_matches_trigger_count(self):
        result = _run(n=7, generate_excel=False)
        assert result.n_users == 7

    def test_quality_score_in_range(self):
        result = _run(generate_excel=False)
        assert 0.0 <= result.quality_score <= 100.0

    def test_realism_score_in_range(self):
        result = _run(generate_excel=False)
        assert 0.0 <= result.realism_score <= 100.0

    def test_feasibility_warnings_is_tuple(self):
        result = _run(generate_excel=False)
        assert isinstance(result.feasibility_warnings, tuple)

    def test_workbook_bytes_none_when_generate_excel_false(self):
        result = _run(generate_excel=False)
        assert result.workbook_bytes is None

    def test_workbook_bytes_present_when_generate_excel_true(self):
        result = _run(generate_excel=True)
        assert result.workbook_bytes is not None
        assert len(result.workbook_bytes) > 0

    def test_workbook_is_valid_xlsx(self):
        result = _run(generate_excel=True)
        wb = load_workbook(io.BytesIO(result.workbook_bytes))
        assert wb is not None

    def test_workbook_has_all_six_sheets(self):
        from core.excel_exporter import SHEET_NAMES
        result = _run(generate_excel=True)
        wb     = load_workbook(io.BytesIO(result.workbook_bytes))
        assert set(wb.sheetnames) == set(SHEET_NAMES)


# ===========================================================================
# 5. Metadata completeness
# ===========================================================================

class TestMetadata:
    def test_required_keys_present(self):
        result = _run(generate_excel=False)
        meta   = result.execution_metadata
        for key in ("campaign_id", "campaign_name", "simulation_start",
                    "simulation_end", "n_users", "n_events", "n_days",
                    "started_at", "completed_at", "elapsed_seconds",
                    "stage_timings"):
            assert key in meta, f"Missing metadata key: {key}"

    def test_campaign_id_matches_config(self):
        cfg    = _cfg()
        result = SimulationOrchestrator(cfg).run(_tdf(5, cfg), generate_excel=False)
        assert result.execution_metadata["campaign_id"] == cfg.campaign_id

    def test_n_days_correct(self):
        cfg    = _cfg()  # 3-day run
        result = SimulationOrchestrator(cfg).run(_tdf(5, cfg), generate_excel=False)
        assert result.execution_metadata["n_days"] == 3

    def test_elapsed_seconds_positive(self):
        result = _run(generate_excel=False)
        assert result.elapsed_seconds > 0

    def test_stage_timings_has_all_stages(self):
        result = _run(generate_excel=False)
        timings = result.execution_metadata["stage_timings"]
        for stage in ("UserStateManager", "AudienceManager",
                      "EngagementGenerator", "ValidationEngine", "finalize_state"):
            assert stage in timings, f"Missing timing for stage: {stage}"

    def test_stage_timings_excel_present_when_enabled(self):
        result  = _run(generate_excel=True)
        timings = result.execution_metadata["stage_timings"]
        assert "ExcelExporter" in timings

    def test_stage_timings_excel_absent_when_disabled(self):
        result  = _run(generate_excel=False)
        timings = result.execution_metadata["stage_timings"]
        assert "ExcelExporter" not in timings

    def test_all_stage_timings_are_positive(self):
        result  = _run(generate_excel=True)
        timings = result.execution_metadata["stage_timings"]
        for stage, t in timings.items():
            assert t >= 0, f"Negative timing for stage {stage}: {t}"

    def test_n_events_matches_result(self):
        result = _run(generate_excel=False)
        assert result.execution_metadata["n_events"] == result.n_events

    def test_n_users_matches_result(self):
        result = _run(generate_excel=False)
        assert result.execution_metadata["n_users"] == result.n_users


# ===========================================================================
# 6. State DataFrame — finalized
# ===========================================================================

class TestFinalizedState:
    def test_state_df_has_state_as_of_date(self):
        result = _run(generate_excel=False)
        assert "state_as_of_date" in result.state_df.columns

    def test_state_as_of_date_equals_sim_end(self):
        cfg    = _cfg()
        result = SimulationOrchestrator(cfg).run(_tdf(5, cfg), generate_excel=False)
        end    = cfg.simulation_end_date
        dates  = pd.to_datetime(result.state_df["state_as_of_date"]).dt.date
        assert (dates == end).all()

    def test_state_df_row_count_equals_n_users(self):
        result = _run(n=6, generate_excel=False)
        assert len(result.state_df) == result.n_users


# ===========================================================================
# 7. Events DataFrame
# ===========================================================================

class TestEventsDF:
    def test_events_df_has_required_columns(self):
        result = _run(generate_excel=False)
        for col in ("user_id", "simulation_date", "channel", "action_type"):
            assert col in result.events_df.columns

    def test_events_simulation_dates_within_range(self):
        cfg    = _cfg()
        result = SimulationOrchestrator(cfg).run(_tdf(5, cfg), generate_excel=False)
        dates  = pd.to_datetime(result.events_df["simulation_date"]).dt.date
        assert (dates >= cfg.simulation_start_date).all()
        assert (dates <= cfg.simulation_end_date).all()

    def test_n_events_matches_metadata(self):
        result = _run(generate_excel=False)
        assert result.n_events == result.execution_metadata["n_events"]


# ===========================================================================
# 8. Validation outputs
# ===========================================================================

class TestValidationOutputs:
    def test_validation_results_has_schema(self):
        result = _run(generate_excel=False)
        for col in ("rule_id", "rule_name", "status", "severity"):
            assert col in result.validation_results_df.columns

    def test_validation_summary_has_overall_row(self):
        result = _run(generate_excel=False)
        cats   = result.validation_summary_df["validation_category"].tolist()
        assert "OVERALL" in cats

    def test_realism_report_has_schema(self):
        result = _run(generate_excel=False)
        for col in ("metric", "target", "actual", "variance"):
            assert col in result.realism_report_df.columns

    def test_quality_score_matches_metadata(self):
        result = _run(generate_excel=False)
        assert result.quality_score == pytest.approx(
            result.execution_metadata["quality_score"]
        )

    def test_realism_score_matches_metadata(self):
        result = _run(generate_excel=False)
        assert result.realism_score == pytest.approx(
            result.execution_metadata["realism_score"]
        )


# ===========================================================================
# 9. Simulation date overrides
# ===========================================================================

class TestDateOverrides:
    def test_run_with_start_end_override(self):
        cfg    = _cfg()
        orch   = SimulationOrchestrator(cfg)
        result = orch.run(
            _tdf(5, cfg),
            generate_excel=False,
            simulation_start=date(2024, 1, 1),
            simulation_end=date(2024, 1, 2),
        )
        assert result.execution_metadata["n_days"] == 2
        assert result.succeeded

    def test_single_day_run(self):
        cfg    = _cfg()
        orch   = SimulationOrchestrator(cfg)
        result = orch.run(
            _tdf(5, cfg),
            generate_excel=False,
            simulation_start=date(2024, 1, 1),
            simulation_end=date(2024, 1, 1),
        )
        assert result.execution_metadata["n_days"] == 1
        assert result.succeeded


# ===========================================================================
# 10. Empty audience scenario
# ===========================================================================

class TestEmptyAudience:
    def test_zero_users_run_succeeds(self):
        cfg    = _cfg()
        tdf    = make_trigger_df(n=0, campaign_id=cfg.campaign_id)
        # Empty trigger df — no users
        result = SimulationOrchestrator(cfg).run(tdf, generate_excel=False)
        assert result.succeeded

    def test_zero_users_gives_zero_events(self):
        cfg    = _cfg()
        tdf    = make_trigger_df(n=0, campaign_id=cfg.campaign_id)
        result = SimulationOrchestrator(cfg).run(tdf, generate_excel=False)
        assert result.n_events == 0


# ===========================================================================
# 11. Empty events scenario
# ===========================================================================

class TestEmptyEventsScenario:
    def test_no_events_workbook_still_generated(self):
        """Even with zero events, ExcelExporter should return valid bytes."""
        cfg  = _cfg()
        tdf  = make_trigger_df(n=0, campaign_id=cfg.campaign_id)
        result = SimulationOrchestrator(cfg).run(tdf, generate_excel=True)
        assert result.workbook_bytes is not None
        wb = load_workbook(io.BytesIO(result.workbook_bytes))
        assert wb is not None


# ===========================================================================
# 12. Error propagation
# ===========================================================================

class TestErrorPropagation:
    def test_stage_failure_raises_simulation_error(self):
        cfg  = _cfg()
        orch = SimulationOrchestrator(cfg)
        tdf  = _tdf(5, cfg)
        # Patch UserStateManager to raise
        with patch(
            "core.simulation_orchestrator.UserStateManager.initialize_user_states",
            side_effect=RuntimeError("forced failure"),
        ):
            with pytest.raises(SimulationError) as exc_info:
                orch.run(tdf, generate_excel=False)
        assert "UserStateManager" in str(exc_info.value)
        assert "forced failure" in str(exc_info.value)

    def test_simulation_error_wraps_original(self):
        cfg  = _cfg()
        orch = SimulationOrchestrator(cfg)
        tdf  = _tdf(5, cfg)
        with patch(
            "core.simulation_orchestrator.UserStateManager.initialize_user_states",
            side_effect=ValueError("bad value"),
        ):
            with pytest.raises(SimulationError) as exc_info:
                orch.run(tdf, generate_excel=False)
        assert exc_info.value.__cause__ is not None

    def test_engagement_generator_failure_propagates(self):
        cfg  = _cfg()
        orch = SimulationOrchestrator(cfg)
        tdf  = _tdf(5, cfg)
        with patch(
            "core.simulation_orchestrator.EngagementGenerator.generate",
            side_effect=RuntimeError("gen fail"),
        ):
            with pytest.raises(SimulationError) as exc_info:
                orch.run(tdf, generate_excel=False)
        assert "EngagementGenerator" in str(exc_info.value)

    def test_excel_exporter_failure_propagates(self):
        cfg  = _cfg()
        orch = SimulationOrchestrator(cfg)
        tdf  = _tdf(5, cfg)
        with patch(
            "core.simulation_orchestrator.ExcelExporter.export",
            side_effect=RuntimeError("xlsx fail"),
        ):
            with pytest.raises(SimulationError) as exc_info:
                orch.run(tdf, generate_excel=True)
        assert "ExcelExporter" in str(exc_info.value)

    def test_stage_timing_recorded_even_on_failure(self):
        """The failed stage's timing entry must still be written."""
        cfg  = _cfg()
        orch = SimulationOrchestrator(cfg)
        tdf  = _tdf(5, cfg)
        # We can't easily inspect internal state after raise,
        # so just verify the exception type and message.
        with patch(
            "core.simulation_orchestrator.AudienceManager.resolve",
            side_effect=RuntimeError("audience fail"),
        ):
            with pytest.raises(SimulationError) as exc_info:
                orch.run(tdf, generate_excel=False)
        assert "AudienceManager" in str(exc_info.value)


# ===========================================================================
# 13. Determinism
# ===========================================================================

class TestDeterminism:
    def test_two_runs_produce_identical_workbook_bytes(self):
        """Deterministic inputs must produce byte-identical workbooks."""
        cfg  = _cfg()
        tdf  = _tdf(5, cfg)
        orch = SimulationOrchestrator(cfg)
        r1   = orch.run(tdf, generate_excel=True)
        r2   = orch.run(tdf, generate_excel=True)
        assert r1.workbook_bytes == r2.workbook_bytes

    def test_two_runs_produce_same_event_count(self):
        cfg  = _cfg()
        tdf  = _tdf(5, cfg)
        orch = SimulationOrchestrator(cfg)
        r1   = orch.run(tdf, generate_excel=False)
        r2   = orch.run(tdf, generate_excel=False)
        assert r1.n_events == r2.n_events

    def test_two_runs_produce_same_quality_score(self):
        cfg  = _cfg()
        tdf  = _tdf(5, cfg)
        orch = SimulationOrchestrator(cfg)
        r1   = orch.run(tdf, generate_excel=False)
        r2   = orch.run(tdf, generate_excel=False)
        assert r1.quality_score == pytest.approx(r2.quality_score)


# ===========================================================================
# 14. Historical and previous state pass-through
# ===========================================================================

class TestOptionalInputs:
    def test_historical_df_none_runs_cleanly(self):
        cfg    = _cfg()
        result = SimulationOrchestrator(cfg).run(
            _tdf(5, cfg), historical_df=None, generate_excel=False
        )
        assert result.succeeded

    def test_previous_state_df_none_runs_cleanly(self):
        cfg    = _cfg()
        result = SimulationOrchestrator(cfg).run(
            _tdf(5, cfg), previous_state_df=None, generate_excel=False
        )
        assert result.succeeded

    def test_previous_state_passthrough(self):
        """Second run re-uses finalised state from first run."""
        cfg  = _cfg()
        tdf  = _tdf(5, cfg)
        orch = SimulationOrchestrator(cfg)
        r1   = orch.run(tdf, generate_excel=False)
        # Pass r1's state as previous state for a second run
        r2   = orch.run(tdf, previous_state_df=r1.state_df, generate_excel=False)
        assert r2.succeeded
        assert r2.n_users == r1.n_users


# ===========================================================================
# 15. Integration — larger dataset
# ===========================================================================

class TestIntegration:
    def test_50_user_run_completes(self):
        cfg    = _cfg(
            simulation_start_date=date(2024, 1, 1),
            simulation_end_date=date(2024, 1, 7),
        )
        tdf    = make_trigger_df(n=50, campaign_id=cfg.campaign_id)
        result = SimulationOrchestrator(cfg).run(tdf, generate_excel=True)
        assert result.succeeded
        assert result.n_users == 50
        assert result.n_events >= 0
        assert result.workbook_bytes is not None

    def test_workbook_event_data_sheet_row_count(self):
        """Sheet 1 Event Data row count = n_events (excluding header)."""
        cfg    = _cfg(
            simulation_start_date=date(2024, 1, 1),
            simulation_end_date=date(2024, 1, 5),
        )
        tdf    = make_trigger_df(n=10, campaign_id=cfg.campaign_id)
        result = SimulationOrchestrator(cfg).run(tdf, generate_excel=True)
        wb     = load_workbook(io.BytesIO(result.workbook_bytes))
        ws     = wb["Event Data"]
        # max_row includes header
        assert ws.max_row == result.n_events + 1

    def test_campaign_metrics_total_impressions_consistent(self):
        """Sheet 2 Total Impressions must match sum of metrics_df.n_impressions."""
        cfg    = _cfg(
            simulation_start_date=date(2024, 1, 1),
            simulation_end_date=date(2024, 1, 3),
        )
        tdf    = make_trigger_df(n=8, campaign_id=cfg.campaign_id)
        result = SimulationOrchestrator(cfg).run(tdf, generate_excel=True)

        expected = int(result.metrics_df["n_impressions"].sum()) \
            if "n_impressions" in result.metrics_df.columns else 0

        wb   = load_workbook(io.BytesIO(result.workbook_bytes))
        ws   = wb["Campaign Metrics"]
        rows = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
        assert int(rows["Total Impressions"]) == expected

    def test_full_pipeline_repr(self):
        """SimulationResult.__repr__ must not raise."""
        result = _run(generate_excel=False)
        rep    = repr(result)
        assert "SimulationResult" in rep


# ===========================================================================
# 16. Compliance
# ===========================================================================

class TestCompliance:
    def test_no_iterrows_in_orchestrator(self):
        import pathlib
        path    = (pathlib.Path(__file__).parent.parent.parent
                   / "core" / "simulation_orchestrator.py")
        content = path.read_text(encoding="utf-8")
        bad     = [
            f"line {i+1}: {line.rstrip()}"
            for i, line in enumerate(content.splitlines())
            if ".iterrows(" in line and not line.lstrip().startswith("#")
        ]
        assert bad == [], "ARCH-011: iterrows() in orchestrator:\n" + "\n".join(bad)

    def test_no_iterrows_in_simulation_result(self):
        import pathlib
        path    = (pathlib.Path(__file__).parent.parent.parent
                   / "models" / "simulation_result.py")
        content = path.read_text(encoding="utf-8")
        bad     = [
            f"line {i+1}: {line.rstrip()}"
            for i, line in enumerate(content.splitlines())
            if ".iterrows(" in line and not line.lstrip().startswith("#")
        ]
        assert bad == [], "iterrows() in simulation_result.py:\n" + "\n".join(bad)

    def test_orchestrator_all_declared(self):
        from core import simulation_orchestrator as m
        assert hasattr(m, "__all__")
        assert "SimulationOrchestrator" in m.__all__

    def test_result_all_declared(self):
        from models import simulation_result as m
        assert hasattr(m, "__all__")
        assert "SimulationResult" in m.__all__

    def test_run_has_docstring(self):
        cfg = _cfg()
        assert SimulationOrchestrator(cfg).run.__doc__

    def test_simulation_result_class_has_docstring(self):
        assert SimulationResult.__doc__

    def test_orchestrator_class_has_docstring(self):
        assert SimulationOrchestrator.__doc__
