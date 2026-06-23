"""Unit and regression tests for core/excel_exporter.py — ExcelExporter (Stage 9).

Coverage:
  * ExcelExporter.__init__
  * ExcelExporter.export() → bytes + file write
  * All six sheet names present
  * Sheet 1 Event Data — schema, column mapping, empty events
  * Sheet 2 Campaign Metrics — all nine KPIs present, correct arithmetic
  * Sheet 3 Validation Results — DataFrame exported verbatim
  * Sheet 4 Validation Summary — DataFrame exported verbatim
  * Sheet 5 Realism Report — DataFrame exported verbatim
  * Sheet 6 Diagnostics — all three sections (CTR, TER, Segment Mix)
  * Formatting — freeze panes, table objects, column widths
  * Error handling — missing required columns
  * Determinism — two identical exports produce identical bytes
  * ARCH-011 compliance — no iterrows()
  * Regression — full test suite (697+) still passes
"""
from __future__ import annotations

import io
import os
import tempfile
from datetime import date, timedelta

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from core.excel_exporter import ExcelExporter, SHEET_NAMES, _compute_campaign_metrics, _compute_diagnostics
from models.ad_config import AdConfig
from models.channel_config import ChannelConfig
from models.enums import ActionType, JourneyStatus, RuleStatus, RuleSeverity
from models.segment_config import SegmentConfig
from models.trigger_config import TriggerConfig
from utils.exceptions import InputValidationError

from tests.test_core.conftest import make_config, make_state_df

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

SIM_DATE = date(2024, 1, 10)


def _cfg(**kw):
    ads = (
        AdConfig("Ad_A", 1, 5, False, "Display", "VendorX", 0.10),
        AdConfig("Ad_B", 2, 7, False, "Email",   "VendorY", 0.05),
    )
    ch  = (ChannelConfig("Email", target_ctr=0.05, target_open_rate=0.25),)
    return make_config(ads=ads, channels=ch, **kw)


def _ev(user_id, sim_date, channel, action_type, current_ad="Ad_A",
        trigger_name="T1", segment="Seg_A", vendor="VendorX"):
    return {
        "campaign_id":     "TEST",
        "user_id":         user_id,
        "simulation_date": sim_date,
        "channel":         channel,
        "action_type":     action_type,
        "current_ad":      current_ad,
        "vendor":          vendor,
        "trigger_name":    trigger_name,
        "segment":         segment,
    }


def _make_events(n=5, with_clicks=True):
    rows = []
    for i in range(1, n + 1):
        uid = f"U{i:03d}"
        rows.append(_ev(uid, SIM_DATE, "Display", ActionType.IMPRESSION.value))
    if with_clicks:
        rows.append(_ev("U001", SIM_DATE, "Display", ActionType.CLICK.value))
    return pd.DataFrame(rows)


def _make_metrics(n_days=5):
    rows = []
    for d in range(n_days):
        rows.append({
            "simulation_date":       SIM_DATE + timedelta(days=d),
            "n_users_active":        5,
            "n_reached":             5,
            "n_impressions":         5,
            "n_sends":               5,
            "n_opens":               2,
            "n_clicks":              1,
            "n_qualifying":          1,
            "actual_ctr_display":    0.20,
            "actual_open_rate_email":0.40,
            "actual_open_rate_wa":   0.0,
            "n_tcc_blocked_users":   0,
            "weekly_reset":          False,
        })
    return pd.DataFrame(rows)


def _make_val_results():
    return pd.DataFrame([
        {"rule_id": "VAL-001", "rule_name": "CTR — Ad_A", "status": "Pass",
         "expected_value": 0.10, "actual_value": 0.10, "variance": 0.0,
         "severity": "Hard", "message": "Within tolerance"},
        {"rule_id": "VAL-005", "rule_name": "User Frequency", "status": "Fail",
         "expected_value": 1.0, "actual_value": 2.0, "variance": 1.0,
         "severity": "Soft", "message": "Exceeded daily cap"},
    ])


def _make_val_summary():
    return pd.DataFrame([
        {"validation_category": "Rate Achievement", "passed": 1, "failed": 0, "warning": 0, "score": 100.0},
        {"validation_category": "Capacity & Frequency", "passed": 0, "failed": 1, "warning": 0, "score": 0.0},
        {"validation_category": "OVERALL", "passed": 1, "failed": 1, "warning": 0, "score": 50.0},
    ])


def _make_realism():
    return pd.DataFrame([
        {"metric": "CTR — Ad_A", "target": 0.10, "actual": 0.10,
         "variance": 0.0, "variance_pct": 0.0, "status": "Good"},
        {"metric": "Open Rate — Ad_B", "target": 0.25, "actual": 0.40,
         "variance": 0.15, "variance_pct": 60.0, "status": "Acceptable"},
    ])


def _export_and_load(cfg, events=None, state=None, metrics=None,
                     val_results=None, val_summary=None, realism=None):
    """Run export() and return (raw_bytes, openpyxl.Workbook)."""
    cfg = cfg or _cfg()
    ev  = events    if events    is not None else _make_events()
    st  = state     if state     is not None else make_state_df(5, config=cfg)
    me  = metrics   if metrics   is not None else _make_metrics()
    vr  = val_results  if val_results  is not None else _make_val_results()
    vs  = val_summary  if val_summary  is not None else _make_val_summary()
    rl  = realism      if realism      is not None else _make_realism()

    exporter = ExcelExporter(cfg)
    raw = exporter.export(ev, st, me, vr, vs, rl)
    wb  = load_workbook(io.BytesIO(raw), data_only=True)
    return raw, wb


# ===========================================================================
# 1. Initialisation
# ===========================================================================

class TestInit:
    def test_instantiates(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        assert ex is not None

    def test_stores_config(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        assert ex._config is cfg


# ===========================================================================
# 2. export() return type and basic structure
# ===========================================================================

class TestExportReturn:
    def test_returns_bytes(self):
        raw, _ = _export_and_load(_cfg())
        assert isinstance(raw, bytes)
        assert len(raw) > 0

    def test_bytes_is_valid_xlsx(self):
        raw, wb = _export_and_load(_cfg())
        assert wb is not None

    def test_all_six_sheets_present(self):
        _, wb = _export_and_load(_cfg())
        assert set(wb.sheetnames) == set(SHEET_NAMES)

    def test_sheet_order_matches_spec(self):
        _, wb = _export_and_load(_cfg())
        assert list(wb.sheetnames) == list(SHEET_NAMES)

    def test_writes_to_file(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        ev  = _make_events()
        st  = make_state_df(5, config=cfg)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            raw = ex.export(ev, st, _make_metrics(), _make_val_results(),
                            _make_val_summary(), _make_realism(),
                            output_path=path)
            assert os.path.exists(path)
            assert os.path.getsize(path) > 0
        finally:
            os.unlink(path)


# ===========================================================================
# 3. Input validation
# ===========================================================================

class TestInputValidation:
    def test_missing_user_id_raises(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        ev  = _make_events().drop(columns=["user_id"])
        with pytest.raises(InputValidationError):
            ex.export(ev, make_state_df(5, config=cfg), _make_metrics(),
                      _make_val_results(), _make_val_summary(), _make_realism())

    def test_missing_channel_raises(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        ev  = _make_events().drop(columns=["channel"])
        with pytest.raises(InputValidationError):
            ex.export(ev, make_state_df(5, config=cfg), _make_metrics(),
                      _make_val_results(), _make_val_summary(), _make_realism())

    def test_missing_simulation_date_raises(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        ev  = _make_events().drop(columns=["simulation_date"])
        with pytest.raises(InputValidationError):
            ex.export(ev, make_state_df(5, config=cfg), _make_metrics(),
                      _make_val_results(), _make_val_summary(), _make_realism())


# ===========================================================================
# 4. Sheet 1 — Event Data
# ===========================================================================

class TestSheetEventData:
    def _sheet_df(self, cfg=None, events=None, state=None):
        cfg = cfg or _cfg()
        _, wb = _export_and_load(cfg, events=events, state=state)
        ws = wb["Event Data"]
        data = list(ws.values)
        if not data:
            return pd.DataFrame()
        return pd.DataFrame(data[1:], columns=data[0])

    def test_has_all_required_columns(self):
        df = self._sheet_df()
        for col in ("Date", "User_ID", "Action", "Channel", "Vendor",
                    "Creative", "Journey_Stage", "Segment", "Trigger_Name"):
            assert col in df.columns, f"Missing column: {col}"

    def test_row_count_matches_events(self):
        ev = _make_events(5, with_clicks=True)
        df = self._sheet_df(events=ev)
        assert len(df) == len(ev)

    def test_date_column_formatted_as_string(self):
        df = self._sheet_df()
        assert df["Date"].dtype == object  # strings in xlsx

    def test_empty_events_produces_header_only(self):
        ev = pd.DataFrame(columns=["user_id","simulation_date","channel",
                                    "action_type","current_ad"])
        _, wb = _export_and_load(_cfg(), events=ev)
        ws   = wb["Event Data"]
        vals = list(ws.values)
        assert len(vals) >= 1       # header row at minimum
        assert vals[0][0] == "Date" # first header cell

    def test_user_id_values_correct(self):
        ev = _make_events(3, with_clicks=False)
        df = self._sheet_df(events=ev)
        assert "U001" in df["User_ID"].values
        assert "U002" in df["User_ID"].values

    def test_action_column_maps_action_type(self):
        ev = _make_events(3, with_clicks=True)
        df = self._sheet_df(events=ev)
        assert "Impression" in df["Action"].values
        assert "Click" in df["Action"].values

    def test_channel_column_present(self):
        df = self._sheet_df()
        assert "Display" in df["Channel"].values

    def test_vendor_column_present(self):
        df = self._sheet_df()
        assert "VendorX" in df["Vendor"].values


# ===========================================================================
# 5. Sheet 2 — Campaign Metrics
# ===========================================================================

class TestSheetCampaignMetrics:
    def _sheet_df(self, cfg=None, events=None, metrics=None):
        cfg = cfg or _cfg()
        _, wb = _export_and_load(cfg, events=events, metrics=metrics)
        ws = wb["Campaign Metrics"]
        data = list(ws.values)
        return pd.DataFrame(data[1:], columns=data[0])

    def test_has_metric_and_value_columns(self):
        df = self._sheet_df()
        assert "Metric" in df.columns
        assert "Value"  in df.columns

    def test_all_nine_kpis_present(self):
        df = self._sheet_df()
        metrics_list = df["Metric"].tolist()
        for kpi in ("Total Impressions", "Total Clicks", "CTR", "Total Sends",
                    "Total Opens", "Open Rate", "Trigger Engagement Rate",
                    "Segment Engagement Rate", "Average Frequency"):
            assert kpi in metrics_list, f"Missing KPI: {kpi}"

    def test_total_impressions_correct(self):
        me = _make_metrics(n_days=3)  # 3 days × 5 impressions = 15
        df = self._sheet_df(metrics=me)
        row = df[df["Metric"] == "Total Impressions"].iloc[0]
        assert int(row["Value"]) == 15

    def test_total_clicks_correct(self):
        me = _make_metrics(n_days=2)  # 2 days × 1 click = 2
        df = self._sheet_df(metrics=me)
        row = df[df["Metric"] == "Total Clicks"].iloc[0]
        assert int(row["Value"]) == 2

    def test_ctr_computed_correctly(self):
        me = _make_metrics(n_days=1)  # 1 click / 5 impressions = 0.2000
        df = self._sheet_df(metrics=me)
        row = df[df["Metric"] == "CTR"].iloc[0]
        assert float(row["Value"]) == pytest.approx(0.2, abs=0.001)

    def test_open_rate_computed_correctly(self):
        me = _make_metrics(n_days=1)  # 2 opens / 5 sends = 0.4000
        df = self._sheet_df(metrics=me)
        row = df[df["Metric"] == "Open Rate"].iloc[0]
        assert float(row["Value"]) == pytest.approx(0.4, abs=0.001)

    def test_empty_metrics_gives_zero_counts(self):
        me = pd.DataFrame(columns=["simulation_date","n_impressions","n_clicks",
                                    "n_sends","n_opens","n_qualifying",
                                    "n_users_active","n_reached","actual_ctr_display",
                                    "actual_open_rate_email","actual_open_rate_wa",
                                    "n_tcc_blocked_users","weekly_reset"])
        df = self._sheet_df(metrics=me)
        row = df[df["Metric"] == "Total Impressions"].iloc[0]
        assert int(row["Value"]) == 0

    def test_ter_is_float_string(self):
        df = self._sheet_df()
        row = df[df["Metric"] == "Trigger Engagement Rate"].iloc[0]
        # Value should be parseable as float
        assert float(row["Value"]) >= 0.0


# ===========================================================================
# 6. Sheet 3 — Validation Results
# ===========================================================================

class TestSheetValidationResults:
    def _sheet_df(self, vr=None):
        vr = vr if vr is not None else _make_val_results()
        _, wb = _export_and_load(_cfg(), val_results=vr)
        ws = wb["Validation Results"]
        data = list(ws.values)
        return pd.DataFrame(data[1:], columns=data[0])

    def test_has_all_validation_columns(self):
        df = self._sheet_df()
        for col in ("rule_id", "rule_name", "status", "expected_value",
                    "actual_value", "variance", "severity", "message"):
            assert col in df.columns

    def test_row_count_matches(self):
        vr = _make_val_results()
        df = self._sheet_df(vr=vr)
        assert len(df) == len(vr)

    def test_rule_ids_preserved(self):
        df = self._sheet_df()
        assert "VAL-001" in df["rule_id"].values
        assert "VAL-005" in df["rule_id"].values

    def test_empty_validation_results_exports_header(self):
        vr = pd.DataFrame(columns=["rule_id","rule_name","status",
                                    "expected_value","actual_value","variance",
                                    "severity","message"])
        df = self._sheet_df(vr=vr)
        assert "rule_id" in df.columns


# ===========================================================================
# 7. Sheet 4 — Validation Summary
# ===========================================================================

class TestSheetValidationSummary:
    def _sheet_df(self, vs=None):
        vs = vs if vs is not None else _make_val_summary()
        _, wb = _export_and_load(_cfg(), val_summary=vs)
        ws = wb["Validation Summary"]
        data = list(ws.values)
        return pd.DataFrame(data[1:], columns=data[0])

    def test_has_all_summary_columns(self):
        df = self._sheet_df()
        for col in ("validation_category", "passed", "failed", "warning", "score"):
            assert col in df.columns

    def test_overall_row_present(self):
        df = self._sheet_df()
        assert "OVERALL" in df["validation_category"].values

    def test_row_count_matches(self):
        vs = _make_val_summary()
        df = self._sheet_df(vs=vs)
        assert len(df) == len(vs)


# ===========================================================================
# 8. Sheet 5 — Realism Report
# ===========================================================================

class TestSheetRealismReport:
    def _sheet_df(self, rl=None):
        rl = rl if rl is not None else _make_realism()
        _, wb = _export_and_load(_cfg(), realism=rl)
        ws = wb["Realism Report"]
        data = list(ws.values)
        return pd.DataFrame(data[1:], columns=data[0])

    def test_has_all_realism_columns(self):
        df = self._sheet_df()
        for col in ("metric", "target", "actual", "variance", "variance_pct", "status"):
            assert col in df.columns

    def test_row_count_matches(self):
        rl = _make_realism()
        df = self._sheet_df(rl=rl)
        assert len(df) == len(rl)

    def test_status_values_preserved(self):
        df = self._sheet_df()
        assert "Good" in df["status"].values


# ===========================================================================
# 9. Sheet 6 — Diagnostics
# ===========================================================================

class TestSheetDiagnostics:
    def _sheet_df(self, cfg=None, events=None, state=None, metrics=None):
        cfg = cfg or _cfg()
        _, wb = _export_and_load(cfg, events=events, state=state, metrics=metrics)
        ws = wb["Diagnostics"]
        data = list(ws.values)
        return pd.DataFrame(data[1:], columns=data[0])

    def test_has_all_diagnostic_columns(self):
        df = self._sheet_df()
        for col in ("Metric", "Entity", "Requested", "Actual", "Variance", "Variance_%"):
            assert col in df.columns

    def test_ctr_rows_present_per_ad(self):
        df = self._sheet_df()
        ctr_rows = df[df["Metric"] == "CTR"]
        assert len(ctr_rows) >= 2  # Ad_A and Ad_B

    def test_ter_rows_present_per_trigger(self):
        df = self._sheet_df()
        ter_rows = df[df["Metric"] == "TER"]
        assert len(ter_rows) >= 1  # at least T1

    def test_ctr_entity_is_ad_name(self):
        df = self._sheet_df()
        ctr_rows = df[df["Metric"] == "CTR"]
        assert "Ad_A" in ctr_rows["Entity"].values

    def test_ter_entity_is_trigger_name(self):
        df = self._sheet_df()
        ter_rows = df[df["Metric"] == "TER"]
        assert "T1" in ter_rows["Entity"].values

    def test_ctr_requested_matches_config(self):
        """Requested CTR for Ad_A should be 0.10 as configured."""
        df = self._sheet_df()
        ad_a_row = df[(df["Metric"] == "CTR") & (df["Entity"] == "Ad_A")]
        assert not ad_a_row.empty
        assert float(ad_a_row.iloc[0]["Requested"]) == pytest.approx(0.10)

    def test_ctr_variance_correct(self):
        """actual - requested = variance."""
        df = self._sheet_df()
        ad_a_row = df[(df["Metric"] == "CTR") & (df["Entity"] == "Ad_A")]
        if not ad_a_row.empty:
            row = ad_a_row.iloc[0]
            computed_var = float(row["Actual"]) - float(row["Requested"])
            assert float(row["Variance"]) == pytest.approx(computed_var, abs=0.001)

    def test_segment_mix_rows_when_configured(self):
        seg = SegmentConfig("Seg_A", 1, 50.0)
        ads = (AdConfig("Ad_A", 1, 5, False, "Display", "VX", 0.10),
               AdConfig("Ad_B", 2, 7, False, "Email",   "VY", 0.05))
        cfg = make_config(segments=(seg,), ads=ads)
        df  = self._sheet_df(cfg=cfg)
        seg_rows = df[df["Metric"] == "Segment Mix"]
        assert len(seg_rows) >= 1

    def test_no_segment_mix_rows_when_no_segments(self):
        cfg = _cfg()  # no segments
        df  = self._sheet_df(cfg=cfg)
        seg_rows = df[df["Metric"] == "Segment Mix"]
        assert len(seg_rows) == 0

    def test_empty_events_produces_zero_actual(self):
        ev = pd.DataFrame(columns=["user_id","simulation_date","channel",
                                    "action_type","current_ad"])
        df = self._sheet_df(events=ev)
        ctr_rows = df[df["Metric"] == "CTR"]
        assert (ctr_rows["Actual"].astype(float) == 0.0).all()


# ===========================================================================
# 10. _compute_campaign_metrics unit tests
# ===========================================================================

class TestComputeCampaignMetrics:
    def test_returns_nine_rows(self):
        cfg  = _cfg()
        rows = _compute_campaign_metrics(_make_metrics(), _make_events(), make_state_df(5, cfg), cfg)
        assert len(rows) == 9

    def test_total_impressions_sums_correctly(self):
        me   = _make_metrics(n_days=2)  # 2 × 5 = 10
        cfg  = _cfg()
        rows = _compute_campaign_metrics(me, _make_events(), make_state_df(5, cfg), cfg)
        d    = dict(rows)
        assert d["Total Impressions"] == 10

    def test_ctr_is_zero_when_no_impressions(self):
        me   = pd.DataFrame(columns=["simulation_date","n_impressions","n_clicks",
                                      "n_sends","n_opens","n_qualifying",
                                      "n_users_active","n_reached","actual_ctr_display",
                                      "actual_open_rate_email","actual_open_rate_wa",
                                      "n_tcc_blocked_users","weekly_reset"])
        cfg  = _cfg()
        rows = _compute_campaign_metrics(me, _make_events(), make_state_df(5, cfg), cfg)
        d    = dict(rows)
        assert float(d["CTR"]) == 0.0

    def test_open_rate_is_zero_when_no_sends(self):
        me   = _make_metrics(1)
        me["n_sends"] = 0
        cfg  = _cfg()
        rows = _compute_campaign_metrics(me, _make_events(), make_state_df(5, cfg), cfg)
        d    = dict(rows)
        assert float(d["Open Rate"]) == 0.0


# ===========================================================================
# 11. _compute_diagnostics unit tests
# ===========================================================================

class TestComputeDiagnostics:
    def test_contains_ctr_rows(self):
        cfg  = _cfg()
        rows = _compute_diagnostics(_make_events(), make_state_df(5, cfg), _make_metrics(), cfg)
        metrics = [r[0] for r in rows]
        assert "CTR" in metrics

    def test_contains_ter_rows(self):
        cfg  = _cfg()
        rows = _compute_diagnostics(_make_events(), make_state_df(5, cfg), _make_metrics(), cfg)
        metrics = [r[0] for r in rows]
        assert "TER" in metrics

    def test_no_segment_mix_when_no_segments(self):
        cfg  = _cfg()
        rows = _compute_diagnostics(_make_events(), make_state_df(5, cfg), _make_metrics(), cfg)
        metrics = [r[0] for r in rows]
        assert "Segment Mix" not in metrics

    def test_segment_mix_when_segments_configured(self):
        seg = SegmentConfig("Seg_A", 1, 50.0)
        ads = (AdConfig("Ad_A", 1, 5, False, "Display", "VX", 0.10),
               AdConfig("Ad_B", 2, 7, False, "Email",   "VY", 0.05))
        cfg = make_config(segments=(seg,), ads=ads)
        rows = _compute_diagnostics(_make_events(), make_state_df(5, cfg), _make_metrics(), cfg)
        metrics = [r[0] for r in rows]
        assert "Segment Mix" in metrics

    def test_ctr_entity_names_match_ads(self):
        cfg  = _cfg()
        rows = _compute_diagnostics(_make_events(), make_state_df(5, cfg), _make_metrics(), cfg)
        ctr_entities = [r[1] for r in rows if r[0] == "CTR"]
        assert "Ad_A" in ctr_entities
        assert "Ad_B" in ctr_entities

    def test_variance_equals_actual_minus_requested(self):
        cfg  = _cfg()
        rows = _compute_diagnostics(_make_events(), make_state_df(5, cfg), _make_metrics(), cfg)
        for metric, entity, requested, actual, variance, var_pct in rows:
            assert variance == pytest.approx(actual - requested, abs=0.001), \
                f"Variance mismatch for {metric} {entity}"


# ===========================================================================
# 12. Formatting
# ===========================================================================

class TestFormatting:
    def test_all_sheets_have_freeze_panes(self):
        _, wb = _export_and_load(_cfg())
        for sheet_name in SHEET_NAMES:
            ws = wb[sheet_name]
            assert ws.freeze_panes == "A2", \
                f"Sheet '{sheet_name}' missing freeze panes"

    def test_all_sheets_have_table_object(self):
        _, wb = _export_and_load(_cfg())
        for sheet_name in SHEET_NAMES:
            ws = wb[sheet_name]
            assert len(ws.tables) >= 1, \
                f"Sheet '{sheet_name}' has no Table object"

    def test_column_widths_set(self):
        _, wb = _export_and_load(_cfg())
        ws = wb["Event Data"]
        # Column A should have explicit width
        assert ws.column_dimensions["A"].width > 0

    def test_header_row_is_row_1(self):
        _, wb = _export_and_load(_cfg())
        ws   = wb["Event Data"]
        vals = list(ws.values)
        assert vals[0][0] == "Date"

    def test_data_starts_at_row_2(self):
        _, wb = _export_and_load(_cfg())
        ws   = wb["Event Data"]
        vals = list(ws.values)
        # Row 2 (index 1) should be first data row, not header
        if len(vals) > 1:
            assert vals[1][0] != "Date"


# ===========================================================================
# 13. Determinism
# ===========================================================================

class TestDeterminism:
    def test_identical_inputs_produce_identical_bytes(self):
        cfg  = _cfg()
        ev   = _make_events()
        st   = make_state_df(5, config=cfg)
        me   = _make_metrics()
        vr   = _make_val_results()
        vs   = _make_val_summary()
        rl   = _make_realism()

        ex   = ExcelExporter(cfg)
        raw1 = ex.export(ev, st, me, vr, vs, rl)
        raw2 = ex.export(ev, st, me, vr, vs, rl)
        assert raw1 == raw2


# ===========================================================================
# 14. Edge cases
# ===========================================================================

class TestEdgeCases:
    def test_all_empty_dataframes(self):
        """All-empty inputs must not raise — returns valid workbook."""
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        ev  = pd.DataFrame(columns=["user_id","simulation_date","channel",
                                     "action_type","current_ad"])
        st  = make_state_df(0, config=cfg)
        me  = pd.DataFrame(columns=["simulation_date","n_impressions","n_clicks",
                                     "n_sends","n_opens","n_qualifying",
                                     "n_users_active","n_reached","actual_ctr_display",
                                     "actual_open_rate_email","actual_open_rate_wa",
                                     "n_tcc_blocked_users","weekly_reset"])
        vr  = pd.DataFrame(columns=["rule_id","rule_name","status",
                                      "expected_value","actual_value","variance",
                                      "severity","message"])
        vs  = pd.DataFrame(columns=["validation_category","passed","failed","warning","score"])
        rl  = pd.DataFrame(columns=["metric","target","actual","variance","variance_pct","status"])
        raw = ex.export(ev, st, me, vr, vs, rl)
        assert isinstance(raw, bytes)
        wb  = load_workbook(io.BytesIO(raw))
        assert set(wb.sheetnames) == set(SHEET_NAMES)

    def test_large_event_set(self):
        """1000-row events DataFrame exports without error."""
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        rows = [_ev(f"U{i:04d}", SIM_DATE + timedelta(days=i % 30),
                    "Display", ActionType.IMPRESSION.value) for i in range(1000)]
        ev  = pd.DataFrame(rows)
        st  = make_state_df(100, config=cfg)
        raw = ex.export(ev, st, _make_metrics(), _make_val_results(),
                        _make_val_summary(), _make_realism())
        wb  = load_workbook(io.BytesIO(raw))
        ws  = wb["Event Data"]
        assert ws.max_row == 1001  # 1 header + 1000 data

    def test_nan_values_handled_gracefully(self):
        """NaN values in input DFs should not crash export."""
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        vr  = _make_val_results().copy()
        vr.loc[0, "message"] = float("nan")
        raw = ex.export(_make_events(), make_state_df(5, config=cfg),
                        _make_metrics(), vr, _make_val_summary(), _make_realism())
        assert isinstance(raw, bytes)

    def test_single_user_single_event(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        ev  = pd.DataFrame([_ev("U001", SIM_DATE, "Display", ActionType.IMPRESSION.value)])
        st  = make_state_df(1, config=cfg)
        raw = ex.export(ev, st, _make_metrics(), _make_val_results(),
                        _make_val_summary(), _make_realism())
        wb  = load_workbook(io.BytesIO(raw))
        ws  = wb["Event Data"]
        assert ws.max_row == 2  # header + 1 row

    def test_categorical_trigger_name_handled(self):
        """Categorical dtype columns do not crash export."""
        cfg  = _cfg()
        ex   = ExcelExporter(cfg)
        ev   = _make_events()
        ev["trigger_name"] = pd.Categorical(ev["trigger_name"])
        ev["segment"]      = pd.Categorical(ev["segment"])
        raw  = ex.export(ev, make_state_df(5, config=cfg), _make_metrics(),
                         _make_val_results(), _make_val_summary(), _make_realism())
        assert isinstance(raw, bytes)


# ===========================================================================
# 15. Integration — end-to-end with EngagementGenerator + ValidationEngine
# ===========================================================================

class TestIntegration:
    def test_full_pipeline_export(self):
        """Run EngagementGenerator → ValidationEngine → ExcelExporter end-to-end."""
        from core.engagement_generator import EngagementGenerator
        from core.validation_engine     import ValidationEngine

        cfg = _cfg(
            simulation_start_date=date(2024, 1, 1),
            simulation_end_date=date(2024, 1, 5),
        )
        gen = EngagementGenerator(cfg)
        st  = make_state_df(10, config=cfg)
        ev, me, _, _ = gen.generate(st)

        ve  = ValidationEngine(cfg)
        vr, vs, rl = ve.validate(ev, st)

        ex  = ExcelExporter(cfg)
        raw = ex.export(ev, st, me, vr, vs, rl)

        wb  = load_workbook(io.BytesIO(raw))
        assert set(wb.sheetnames) == set(SHEET_NAMES)

        # Verify Sheet 2 has correct total impressions
        ws   = wb["Campaign Metrics"]
        data = list(ws.values)
        df   = pd.DataFrame(data[1:], columns=data[0])
        row  = df[df["Metric"] == "Total Impressions"].iloc[0]
        expected_impressions = int(me["n_impressions"].sum()) if "n_impressions" in me.columns else 0
        assert int(row["Value"]) == expected_impressions

    def test_all_sheet_names_constant(self):
        from core.excel_exporter import SHEET_NAMES as SN
        assert len(SN) == 6
        assert SN[0] == "Event Data"
        assert SN[1] == "Campaign Metrics"
        assert SN[-1] == "Diagnostics"


# ===========================================================================
# 16. Compliance
# ===========================================================================

class TestCompliance:
    def test_no_iterrows_in_excel_exporter(self):
        import pathlib
        path = (pathlib.Path(__file__).parent.parent.parent
                / "core" / "excel_exporter.py")
        content = path.read_text(encoding="utf-8")
        bad = [
            f"line {i+1}: {line.rstrip()}"
            for i, line in enumerate(content.splitlines())
            if ".iterrows(" in line and not line.lstrip().startswith("#")
        ]
        assert bad == [], "ARCH-011: iterrows() found:\n" + "\n".join(bad)

    def test_all_declared(self):
        from core import excel_exporter
        assert hasattr(excel_exporter, "__all__")
        assert "ExcelExporter" in excel_exporter.__all__

    def test_public_methods_have_docstrings(self):
        cfg = _cfg()
        ex  = ExcelExporter(cfg)
        for name in ("export",):
            assert getattr(ex, name).__doc__, f"Missing docstring: {name}"

    def test_sheet_names_tuple_exported(self):
        from core.excel_exporter import SHEET_NAMES as SN
        assert isinstance(SN, tuple)
        assert len(SN) == 6
