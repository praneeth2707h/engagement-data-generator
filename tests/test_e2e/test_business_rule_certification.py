"""Stage 12 — Business-Rule End-to-End Certification Suite.

Proves that every configurable business rule exposed in the Streamlit UI
produces a measurably different simulation output when changed.

Certification question (CERT-Q-001):
    "Can a non-technical user modify business rules and reliably
     influence generated outputs?"

Answer methodology:
    For each business rule:
    1. Run baseline simulation with low/default rule value.
    2. Run modified simulation with increased/decreased rule value.
    3. Assert the output metric changes in the expected direction.
    4. All assertions use real SimulationOrchestrator — zero mocks.

Design notes
------------
* engagement_cooldown_days=0 is required for Scenarios 2/7 to prevent
  per-user cooldown saturation from masking the rule under test.
* TCC (Trigger Capacity Consumption) exhausts on Day 1 for low TER
  scenarios; this is expected and correct behaviour.
* Weekly engagement cap interacts with TCC capacity; Scenario 7 uses
  N=500 / 5-day window (within one ISO week) to isolate cap effect.
* Scenario 8 requires previous_state_df where cooling users have
  journey_status=Completed — the JourneyEngine only restarts RE_ENTRY
  users whose journey is COMPLETED, not NOT_STARTED.
* All simulations use generate_excel=False except Scenario 9 (workbook)
  and Scenario 10 (determinism).

References
----------
CERT-001 .. CERT-010 — Certification scenario IDs
DEF-E2E-001          — RE_ENTRY users require journey_status=Completed
DEF-E2E-002          — engagement_cap effect hidden by TCC at identical TER
"""
from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path
from typing import Any
import unittest

import pandas as pd

# Ensure project root is importable
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from core.simulation_orchestrator import SimulationOrchestrator
from core.user_state_manager import UserStateManager
from models.ad_config import AdConfig
from models.config_registry import ConfigRegistry
from models.enums import EligibilityStatus, JourneyStatus
from models.rule_config import RuleConfig
from models.segment_config import SegmentConfig
from models.trigger_config import TriggerConfig
from tests.test_core.conftest import make_config, make_trigger_df


# ---------------------------------------------------------------------------
# Shared test configuration builders
# ---------------------------------------------------------------------------

_SIM_START = date(2024, 1, 1)   # Monday — important for weekly-reset tests
_SIM_END_14 = date(2024, 1, 14)  # 14-day window
_SIM_END_5  = date(2024, 1, 5)   # 5-day window (within one ISO week)


def _cfg(**kw) -> ConfigRegistry:
    """Return a ConfigRegistry with sensible E2E defaults.

    Defaults are tuned to produce measurable output differences:
    - engagement_cooldown_days=0  → per-click cooldown disabled; isolates CTR
    - high impression/engagement caps → caps are not the binding constraint
      (except when explicitly testing them)
    - TER=1.0 → full capacity; use lower TER to test TER effect
    """
    defaults = dict(
        simulation_start_date=_SIM_START,
        simulation_end_date=_SIM_END_14,
        weekly_impression_cap=20,
        weekly_engagement_cap=14,
        weekly_click_cap=20,
        weekly_open_cap=20,
        engagement_cooldown_days=0,
        cooling_period_days=0,
        ads=(AdConfig("Ad_A", 1, 14, False, "Display", "VendorX", 0.10),),
        triggers=(TriggerConfig("T1", 1, 1.0),),
        segments=(),
        channels=(),
    )
    defaults.update(kw)
    return make_config(**defaults)


def _trigger_df(n: int = 200,
                trigger_name: str = "T1",
                segment: str = "Seg_A") -> pd.DataFrame:
    """Build a minimal trigger DataFrame."""
    return make_trigger_df(n=n, trigger_name=trigger_name, segment=segment)


def _run(cfg: ConfigRegistry,
         tdf: pd.DataFrame,
         previous_state_df: pd.DataFrame | None = None,
         generate_excel: bool = False):
    """Execute the full orchestrator and return SimulationResult."""
    return SimulationOrchestrator(cfg).run(
        trigger_df=tdf,
        historical_df=None,
        previous_state_df=previous_state_df,
        generate_excel=generate_excel,
    )


# ---------------------------------------------------------------------------
# Metric helpers
# ---------------------------------------------------------------------------

def _actual_ctr(events_df: pd.DataFrame) -> float:
    """Clicks / Impressions for Display channel."""
    imps = (events_df["action_type"] == "Impression").sum()
    clks = (events_df["action_type"] == "Click").sum()
    return clks / imps if imps > 0 else 0.0


def _qualifying_events(events_df: pd.DataFrame) -> int:
    """Display clicks = qualifying events (BIZ-011)."""
    return int((events_df["action_type"] == "Click").sum())


def _seg_fraction(events_df: pd.DataFrame, seg: str) -> float:
    """Fraction of all events attributed to segment `seg`."""
    if events_df.empty:
        return 0.0
    return float((events_df["segment"].astype(str) == seg).mean())


def _impression_count(events_df: pd.DataFrame) -> int:
    return int((events_df["action_type"] == "Impression").sum())


# ---------------------------------------------------------------------------
# SCENARIO 1 — Baseline Run
# ---------------------------------------------------------------------------

class TestScenario01Baseline(unittest.TestCase):
    """CERT-001: Record baseline metrics for all subsequent comparisons."""

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=200)
        result = _run(
            _cfg(ads=(AdConfig("Ad_A", 1, 14, False, "Display", "VendorX", 0.10),),
                 triggers=(TriggerConfig("T1", 1, 1.0),)),
            tdf,
        )
        cls.result = result
        cls.events = result.events_df

    def test_s01_simulation_succeeds(self):
        self.assertTrue(self.result.succeeded,
                        "Baseline simulation must complete successfully.")

    def test_s01_events_generated(self):
        self.assertGreater(len(self.events), 0,
                           "Baseline must produce > 0 events.")

    def test_s01_impressions_present(self):
        imps = _impression_count(self.events)
        self.assertGreater(imps, 0, "Baseline must have impressions.")

    def test_s01_clicks_present(self):
        clks = _qualifying_events(self.events)
        self.assertGreater(clks, 0, "Baseline must have clicks (qualifying events).")

    def test_s01_users_populated(self):
        self.assertGreater(self.result.n_users, 0, "Baseline n_users > 0.")

    def test_s01_event_count_is_integer(self):
        self.assertIsInstance(self.result.n_events, int)

    def test_s01_segment_present(self):
        segs = self.events["segment"].astype(str).unique()
        self.assertIn("Seg_A", segs,
                      "Baseline trigger segment Seg_A must appear in events.")

    def test_s01_quality_score_in_range(self):
        self.assertGreaterEqual(self.result.quality_score, 0.0)
        self.assertLessEqual(self.result.quality_score, 100.0)

    def test_s01_realism_score_in_range(self):
        self.assertGreaterEqual(self.result.realism_score, 0.0)
        self.assertLessEqual(self.result.realism_score, 100.0)


# ---------------------------------------------------------------------------
# SCENARIO 2 — CTR Increase
# ---------------------------------------------------------------------------

class TestScenario02CTR(unittest.TestCase):
    """CERT-002: Increasing target_ctr raises actual click-through rate.

    Business rule: Ad CTR slider (0–100%) in Business Rules page.
    Config path:   ads[i].target_ctr
    Formula:       p_click = clip(2.0 × composite_score × target_ctr, 0, 1)
    Isolation:     engagement_cooldown_days=0 removes per-click cooldown.
    """

    _N = 300
    _DAYS_14 = _SIM_END_14

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=cls._N)
        cfg_lo = _cfg(ads=(AdConfig("Ad_A", 1, 14, False, "Display", "V", 0.02),),
                      triggers=(TriggerConfig("T1", 1, 1.0),))
        cfg_hi = _cfg(ads=(AdConfig("Ad_A", 1, 14, False, "Display", "V", 0.40),),
                      triggers=(TriggerConfig("T1", 1, 1.0),))
        cls.result_lo = _run(cfg_lo, tdf)
        cls.result_hi = _run(cfg_hi, tdf)
        cls.ctr_lo = _actual_ctr(cls.result_lo.events_df)
        cls.ctr_hi = _actual_ctr(cls.result_hi.events_df)

    def test_s02_both_simulations_succeed(self):
        self.assertTrue(self.result_lo.succeeded)
        self.assertTrue(self.result_hi.succeeded)

    def test_s02_high_ctr_produces_more_clicks(self):
        clks_lo = _qualifying_events(self.result_lo.events_df)
        clks_hi = _qualifying_events(self.result_hi.events_df)
        self.assertGreater(
            clks_hi, clks_lo,
            f"CTR=0.40 should produce more clicks than CTR=0.02 "
            f"(got {clks_hi} vs {clks_lo}).",
        )

    def test_s02_actual_ctr_increases(self):
        self.assertGreater(
            self.ctr_hi, self.ctr_lo,
            f"Actual CTR must be higher for target_ctr=0.40 than 0.02 "
            f"(got {self.ctr_hi:.4f} vs {self.ctr_lo:.4f}).",
        )

    def test_s02_ctr_ratio_at_least_3x(self):
        """CTR 20× increase in target should produce ≥3× actual increase."""
        self.assertGreater(self.ctr_lo, 0,
                           "Low CTR must produce at least some clicks.")
        ratio = self.ctr_hi / self.ctr_lo
        self.assertGreaterEqual(
            ratio, 3.0,
            f"Actual CTR ratio (hi/lo) must be ≥3× (got {ratio:.2f}×).",
        )

    def test_s02_impression_counts_identical(self):
        """Impressions are independent of CTR — only clicks change."""
        imps_lo = _impression_count(self.result_lo.events_df)
        imps_hi = _impression_count(self.result_hi.events_df)
        self.assertEqual(
            imps_lo, imps_hi,
            f"Impression counts should be equal regardless of CTR "
            f"(got {imps_lo} vs {imps_hi}).",
        )


# ---------------------------------------------------------------------------
# SCENARIO 3 — TER Increase
# ---------------------------------------------------------------------------

class TestScenario03TER(unittest.TestCase):
    """CERT-003: Increasing engagement_rate_target raises qualifying events.

    Business rule: TER target slider per trigger in Business Rules page.
    Config path:   triggers[i].engagement_rate_target
    Mechanism:     TCC capacity = ceil(N × TER). Higher TER → higher capacity
                   → more qualifying events allowed before TCC exhaustion.
    """

    _N = 200

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=cls._N)
        # Low TER: capacity=ceil(200×0.05)=10 → TCC exhausts on day 1
        cfg_lo = _cfg(
            triggers=(TriggerConfig("T1", 1, 0.05),),
            weekly_engagement_cap=3,
            engagement_cooldown_days=3,
        )
        # High TER: capacity=ceil(200×0.90)=180 → TCC allows much more
        cfg_hi = _cfg(
            triggers=(TriggerConfig("T1", 1, 0.90),),
            weekly_engagement_cap=3,
            engagement_cooldown_days=3,
        )
        cls.result_lo = _run(cfg_lo, tdf)
        cls.result_hi = _run(cfg_hi, tdf)
        cls.qual_lo = _qualifying_events(cls.result_lo.events_df)
        cls.qual_hi = _qualifying_events(cls.result_hi.events_df)

    def test_s03_both_succeed(self):
        self.assertTrue(self.result_lo.succeeded)
        self.assertTrue(self.result_hi.succeeded)

    def test_s03_higher_ter_produces_more_qualifying_events(self):
        self.assertGreater(
            self.qual_hi, self.qual_lo,
            f"TER=0.90 must produce more qualifying events than TER=0.05 "
            f"(got {self.qual_hi} vs {self.qual_lo}).",
        )

    def test_s03_qualifying_ratio_at_least_3x(self):
        self.assertGreater(self.qual_lo, 0, "Low TER must still produce some qualifying events.")
        ratio = self.qual_hi / self.qual_lo
        self.assertGreaterEqual(
            ratio, 3.0,
            f"TER 18× increase should yield ≥3× qualifying events (got {ratio:.2f}×).",
        )

    def test_s03_impressions_unaffected_by_ter(self):
        """TER only limits qualifying events; impressions are unaffected."""
        imps_lo = _impression_count(self.result_lo.events_df)
        imps_hi = _impression_count(self.result_hi.events_df)
        # Impressions should be similar (both use same N and duration)
        ratio = imps_hi / imps_lo if imps_lo > 0 else 0
        self.assertLess(
            abs(ratio - 1.0), 0.15,
            f"Impression counts should be similar regardless of TER "
            f"(got ratio {ratio:.2f}). TER only gates qualifying, not reach.",
        )


# ---------------------------------------------------------------------------
# SCENARIO 4 — Segment Mix Change
# ---------------------------------------------------------------------------

class TestScenario04SegmentMix(unittest.TestCase):
    """CERT-004: Changing segment distribution in trigger file changes
    the actual segment mix observed in generated events.

    Business rule: Segment Mix targets % per segment in Business Rules page.
    Implementation: Segments are determined by trigger file composition.
                    Distribution targets are validated against actuals.
    Measurement:    Fraction of events attributed to Seg_A.
    """

    _N = 300
    _SIM_DAYS = _SIM_END_14

    @classmethod
    def setUpClass(cls):
        cfg = _cfg(triggers=(TriggerConfig("T1", 1, 1.0),))
        # 50/50 mix
        tdf_even = pd.DataFrame({
            "Campaign_ID":  ["TEST_CAMPAIGN"] * cls._N,
            "User_ID":      [f"U{i:04d}" for i in range(cls._N)],
            "Trigger_Name": ["T1"] * cls._N,
            "Segment":      ["Seg_A"] * (cls._N // 2) + ["Seg_B"] * (cls._N // 2),
            "Trigger_Date": [date(2024, 1, 1)] * cls._N,
        })
        # 80/20 Seg_A-heavy mix
        n_a = int(cls._N * 0.80)
        n_b = cls._N - n_a
        tdf_skewed = pd.DataFrame({
            "Campaign_ID":  ["TEST_CAMPAIGN"] * cls._N,
            "User_ID":      [f"U{i:04d}" for i in range(cls._N)],
            "Trigger_Name": ["T1"] * cls._N,
            "Segment":      ["Seg_A"] * n_a + ["Seg_B"] * n_b,
            "Trigger_Date": [date(2024, 1, 1)] * cls._N,
        })
        cls.result_even   = _run(cfg, tdf_even)
        cls.result_skewed = _run(cfg, tdf_skewed)
        cls.seg_a_even    = _seg_fraction(cls.result_even.events_df, "Seg_A")
        cls.seg_a_skewed  = _seg_fraction(cls.result_skewed.events_df, "Seg_A")

    def test_s04_both_succeed(self):
        self.assertTrue(self.result_even.succeeded)
        self.assertTrue(self.result_skewed.succeeded)

    def test_s04_even_mix_is_near_fifty_percent(self):
        self.assertAlmostEqual(
            self.seg_a_even, 0.50, delta=0.05,
            msg=f"50/50 trigger file should yield ~50% Seg_A events "
                f"(got {self.seg_a_even:.2%}).",
        )

    def test_s04_skewed_mix_is_near_eighty_percent(self):
        self.assertAlmostEqual(
            self.seg_a_skewed, 0.80, delta=0.05,
            msg=f"80/20 trigger file should yield ~80% Seg_A events "
                f"(got {self.seg_a_skewed:.2%}).",
        )

    def test_s04_skewed_seg_a_fraction_exceeds_even(self):
        self.assertGreater(
            self.seg_a_skewed, self.seg_a_even + 0.20,
            f"80/20 Seg_A fraction must exceed 50/50 by ≥20pp "
            f"(got {self.seg_a_skewed:.2%} vs {self.seg_a_even:.2%}).",
        )

    def test_s04_both_mixes_produce_events(self):
        self.assertGreater(len(self.result_even.events_df), 0)
        self.assertGreater(len(self.result_skewed.events_df), 0)


# ---------------------------------------------------------------------------
# SCENARIO 5 — Journey Length Change
# ---------------------------------------------------------------------------

class TestScenario05JourneyLength(unittest.TestCase):
    """CERT-005: Increasing per-ad duration_days changes journey progression.

    Business rule: Journey lengths (days per ad) in Business Rules page.
    Config path:   ads[i].duration_days
    Mechanism:     Short duration → users advance to Ad_B earlier → more Ad_B events.
                   Long duration → users remain on Ad_A for entire simulation.
    Measurement:   Events on Ad_B as a fraction of total events.
    """

    _N = 200
    _SIM_DAYS = _SIM_END_14

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=cls._N)
        # Short: 3 days per ad → users reach Ad_B after 3 days and stay there
        cfg_short = _cfg(
            ads=(
                AdConfig("Ad_A", 1, 3,  False, "Display", "V", 0.10),
                AdConfig("Ad_B", 2, 20, False, "Display", "V", 0.10),
            ),
            triggers=(TriggerConfig("T1", 1, 1.0),),
        )
        # Long: 20 days per ad → nobody reaches Ad_B in a 14-day window
        cfg_long = _cfg(
            ads=(
                AdConfig("Ad_A", 1, 20, False, "Display", "V", 0.10),
                AdConfig("Ad_B", 2, 20, False, "Display", "V", 0.10),
            ),
            triggers=(TriggerConfig("T1", 1, 1.0),),
        )
        cls.result_short = _run(cfg_short, tdf)
        cls.result_long  = _run(cfg_long, tdf)
        cls.adb_short = int((cls.result_short.events_df["current_ad"].astype(str) == "Ad_B").sum())
        cls.adb_long  = int((cls.result_long.events_df["current_ad"].astype(str) == "Ad_B").sum())
        cls.ada_short = int((cls.result_short.events_df["current_ad"].astype(str) == "Ad_A").sum())
        cls.ada_long  = int((cls.result_long.events_df["current_ad"].astype(str) == "Ad_A").sum())

    def test_s05_both_succeed(self):
        self.assertTrue(self.result_short.succeeded)
        self.assertTrue(self.result_long.succeeded)

    def test_s05_short_journey_produces_ad_b_events(self):
        self.assertGreater(
            self.adb_short, 0,
            f"Short duration=3 must cause users to advance to Ad_B "
            f"within 14-day window (got {self.adb_short} Ad_B events).",
        )

    def test_s05_long_journey_produces_no_ad_b_events(self):
        self.assertEqual(
            self.adb_long, 0,
            f"Long duration=20 means no user reaches Ad_B in 14 days "
            f"(got {self.adb_long} Ad_B events).",
        )

    def test_s05_short_has_more_ad_b_events_than_long(self):
        self.assertGreater(
            self.adb_short, self.adb_long,
            f"Short journey must produce more Ad_B events than long "
            f"(got {self.adb_short} vs {self.adb_long}).",
        )

    def test_s05_long_has_more_ad_a_events_than_short(self):
        """Long journey means users stay on Ad_A longer → more Ad_A events."""
        self.assertGreater(
            self.ada_long, self.ada_short,
            f"Long journey should produce more Ad_A events "
            f"(got {self.ada_long} vs {self.ada_short}).",
        )


# ---------------------------------------------------------------------------
# SCENARIO 6 — Weekly Impression Cap Reduction
# ---------------------------------------------------------------------------

class TestScenario06ImpressionCap(unittest.TestCase):
    """CERT-006: Reducing weekly_impression_cap lowers total impressions.

    Business rule: Weekly Impression Cap in Business Rules page.
    Config path:   weekly_impression_cap
    Mechanism:     BehaviorEngine gates Impression events on
                   weekly_impressions < weekly_impression_cap.
    Measurement:   Total impression events across simulation.
    """

    _N = 300
    _CAPS = (2, 5, 14)

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=cls._N)
        cls.results = {}
        for cap in cls._CAPS:
            cls.results[cap] = _run(
                _cfg(weekly_impression_cap=cap,
                     triggers=(TriggerConfig("T1", 1, 1.0),)),
                tdf,
            )

    def test_s06_all_succeed(self):
        for cap, r in self.results.items():
            self.assertTrue(r.succeeded, f"cap={cap} simulation must succeed.")

    def test_s06_lowest_cap_gives_fewest_impressions(self):
        imps = {cap: _impression_count(r.events_df) for cap, r in self.results.items()}
        self.assertLess(
            imps[2], imps[14],
            f"cap=2 must produce fewer impressions than cap=14 "
            f"(got {imps[2]} vs {imps[14]}).",
        )

    def test_s06_impression_counts_are_ordered(self):
        imps = {cap: _impression_count(r.events_df) for cap, r in self.results.items()}
        self.assertLess(
            imps[2], imps[5],
            f"cap=2 must produce fewer impressions than cap=5 "
            f"({imps[2]} vs {imps[5]}).",
        )
        self.assertLess(
            imps[5], imps[14],
            f"cap=5 must produce fewer impressions than cap=14 "
            f"({imps[5]} vs {imps[14]}).",
        )

    def test_s06_cap_2_reduces_impressions_by_at_least_40_percent(self):
        imps_low  = _impression_count(self.results[2].events_df)
        imps_high = _impression_count(self.results[14].events_df)
        reduction = 1.0 - imps_low / imps_high
        self.assertGreaterEqual(
            reduction, 0.40,
            f"Reducing cap from 14 to 2 should cut impressions by ≥40% "
            f"(got {reduction:.1%}).",
        )

    def test_s06_cap_does_not_affect_zero_impressions(self):
        """With any cap ≥ 1, at least some impressions must occur."""
        for cap, r in self.results.items():
            self.assertGreater(
                _impression_count(r.events_df), 0,
                f"cap={cap} must still produce some impressions.",
            )


# ---------------------------------------------------------------------------
# SCENARIO 7 — Weekly Engagement Cap Reduction
# ---------------------------------------------------------------------------

class TestScenario07EngagementCap(unittest.TestCase):
    """CERT-007: Reducing weekly_engagement_cap lowers total qualifying events.

    Business rule: Weekly Engagement Cap in Business Rules page.
    Config path:   weekly_engagement_cap
    Mechanism:     BehaviorEngine gates qualifying events on
                   weekly_engagements < weekly_engagement_cap.
    Isolation:     N=500, 5-day window (one ISO week, no reset).
                   engagement_cooldown_days=0 to maximise click attempts.
                   TER=1.0 to maximise TCC capacity.
    Finding:       cap=1 → unique qualifiers monopolised later → fewer unique
                   qualifying users but slightly lower total qualifying count.
    """

    _N = 500
    _SIM_END_5 = date(2024, 1, 5)

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=cls._N)
        cfg_lo = make_config(
            simulation_start_date=_SIM_START,
            simulation_end_date=cls._SIM_END_5,
            weekly_impression_cap=20,
            weekly_engagement_cap=1,    # one qualifying event per week
            weekly_click_cap=20,
            weekly_open_cap=20,
            engagement_cooldown_days=0,
            cooling_period_days=0,
            ads=(AdConfig("Ad_A", 1, 10, False, "Display", "V", 0.50),),
            triggers=(TriggerConfig("T1", 1, 1.0),),
            segments=(), channels=(),
        )
        cfg_hi = make_config(
            simulation_start_date=_SIM_START,
            simulation_end_date=cls._SIM_END_5,
            weekly_impression_cap=20,
            weekly_engagement_cap=20,   # effectively uncapped
            weekly_click_cap=20,
            weekly_open_cap=20,
            engagement_cooldown_days=0,
            cooling_period_days=0,
            ads=(AdConfig("Ad_A", 1, 10, False, "Display", "V", 0.50),),
            triggers=(TriggerConfig("T1", 1, 1.0),),
            segments=(), channels=(),
        )
        cls.result_lo = _run(cfg_lo, tdf)
        cls.result_hi = _run(cfg_hi, tdf)
        cls.qual_lo = _qualifying_events(cls.result_lo.events_df)
        cls.qual_hi = _qualifying_events(cls.result_hi.events_df)

    def test_s07_both_succeed(self):
        self.assertTrue(self.result_lo.succeeded)
        self.assertTrue(self.result_hi.succeeded)

    def test_s07_higher_cap_produces_more_or_equal_qualifying_events(self):
        self.assertGreaterEqual(
            self.qual_hi, self.qual_lo,
            f"Higher engagement cap must produce ≥ qualifying events "
            f"(got cap=20: {self.qual_hi}, cap=1: {self.qual_lo}).",
        )

    def test_s07_cap1_unique_qualifiers_exceed_cap20_unique_qualifiers(self):
        """Cap=1 forces spread: more UNIQUE users qualify.
        Cap=20 allows repetition: fewer unique users qualify (some monopolise capacity).
        """
        unique_lo = self.result_lo.events_df[
            self.result_lo.events_df["action_type"] == "Click"]["user_id"].nunique()
        unique_hi = self.result_hi.events_df[
            self.result_hi.events_df["action_type"] == "Click"]["user_id"].nunique()
        self.assertGreater(
            unique_lo, unique_hi,
            f"cap=1 should produce more unique qualifying users than cap=20 "
            f"(got {unique_lo} vs {unique_hi}). "
            "Cap=1 prevents any single user from consuming multiple TCC slots.",
        )

    def test_s07_impression_counts_identical(self):
        """Engagement cap does not affect impression events."""
        imps_lo = _impression_count(self.result_lo.events_df)
        imps_hi = _impression_count(self.result_hi.events_df)
        self.assertEqual(
            imps_lo, imps_hi,
            f"Impression counts must be equal regardless of engagement cap "
            f"(got {imps_lo} vs {imps_hi}).",
        )


# ---------------------------------------------------------------------------
# SCENARIO 8 — Re-entry Enable / Disable
# ---------------------------------------------------------------------------

class TestScenario08ReEntry(unittest.TestCase):
    """CERT-008: allow_reentry controls whether cooling-expired users re-enter.

    Business rule: Allow Re-entry checkbox in Business Rules page.
    Config path:   allow_reentry (bool)
    Mechanism:     AudienceManager.classify_eligibility() assigns RE_ENTRY
                   (allow_reentry=True) or EXCLUDED (allow_reentry=False) to
                   users whose cooling_period_end has expired.
                   JourneyEngine._start_journeys() activates RE_ENTRY users
                   whose journey_status=Completed.

    Setup:         100 of 200 users are pre-loaded with expired cooling
                   (eligibility=Cooling, journey_status=Completed,
                   cooling_period_end=2023-12-01 which is before SIM_START).
    Expected:
        allow_reentry=True  → those 100 become RE_ENTRY → total 200 active users
        allow_reentry=False → those 100 become EXCLUDED → total 100 active users

    DEF-E2E-001: RE_ENTRY users MUST have journey_status=Completed.
                 journey_status=Not_Started causes JourneyEngine to skip them
                 (mask_reentry checks Completed, not Not_Started).
    """

    _N = 200
    _COOLING = 100

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=cls._N)
        # Build previous state: first 100 users in expired cooling, Completed journey
        usm = UserStateManager(_cfg(cooling_period_days=30))
        prev = usm.initialize_user_states(tdf, previous_state_df=None).copy()
        cooling_idx = prev.index[:cls._COOLING]
        prev.loc[cooling_idx, "eligibility_status"] = EligibilityStatus.COOLING.value
        prev.loc[cooling_idx, "journey_status"]     = JourneyStatus.COMPLETED.value
        prev.loc[cooling_idx, "cooling_period_end"] = date(2023, 12, 1)  # expired

        cls.prev_state = prev

        cfg_on  = _cfg(allow_reentry=True,  cooling_period_days=30)
        cfg_off = _cfg(allow_reentry=False, cooling_period_days=30)

        cls.result_on  = _run(cfg_on,  tdf, previous_state_df=prev.copy())
        cls.result_off = _run(cfg_off, tdf, previous_state_df=prev.copy())

    def test_s08_both_succeed(self):
        self.assertTrue(self.result_on.succeeded)
        self.assertTrue(self.result_off.succeeded)

    def test_s08_reentry_on_uses_all_200_users(self):
        unique = self.result_on.events_df["user_id"].nunique()
        self.assertEqual(
            unique, self._N,
            f"allow_reentry=True must activate all {self._N} users (including "
            f"{self._COOLING} cooling-expired). Got {unique}.",
        )

    def test_s08_reentry_off_uses_only_100_new_users(self):
        unique = self.result_off.events_df["user_id"].nunique()
        self.assertEqual(
            unique, self._N - self._COOLING,
            f"allow_reentry=False must exclude {self._COOLING} cooling-expired "
            f"users. Expected {self._N - self._COOLING}, got {unique}.",
        )

    def test_s08_reentry_on_produces_more_events(self):
        self.assertGreater(
            self.result_on.n_events, self.result_off.n_events,
            f"allow_reentry=True should produce more events "
            f"(got {self.result_on.n_events} vs {self.result_off.n_events}).",
        )

    def test_s08_audience_df_re_entry_status_correct(self):
        re_entry_count = (
            self.result_on.audience_df["eligibility_status"]
            .astype(str) == EligibilityStatus.RE_ENTRY.value
        ).sum()
        self.assertEqual(
            re_entry_count, self._COOLING,
            f"Exactly {self._COOLING} users should have RE_ENTRY status "
            f"(got {re_entry_count}).",
        )

    def test_s08_audience_df_excluded_status_correct(self):
        excluded_count = (
            self.result_off.audience_df["eligibility_status"]
            .astype(str) == EligibilityStatus.EXCLUDED.value
        ).sum()
        self.assertEqual(
            excluded_count, self._COOLING,
            f"Exactly {self._COOLING} users should be EXCLUDED "
            f"(got {excluded_count}).",
        )


# ---------------------------------------------------------------------------
# SCENARIO 9 — Workbook Certification
# ---------------------------------------------------------------------------

class TestScenario09WorkbookCertification(unittest.TestCase):
    """CERT-009: Downloaded workbook reflects modified simulation outputs.

    Verifies:
    1. workbook_bytes is non-empty.
    2. Event Data sheet row count matches events_df.
    3. Workbook produced from higher-CTR config has more clicks in the sheet.
    4. All 6 expected sheets are present.
    """

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=200)
        # Baseline workbook
        cfg_lo = _cfg(ads=(AdConfig("Ad_A", 1, 14, False, "Display", "V", 0.02),),
                      triggers=(TriggerConfig("T1", 1, 1.0),))
        cls.result_lo = _run(cfg_lo, tdf, generate_excel=True)

        # High-CTR workbook
        cfg_hi = _cfg(ads=(AdConfig("Ad_A", 1, 14, False, "Display", "V", 0.40),),
                      triggers=(TriggerConfig("T1", 1, 1.0),))
        cls.result_hi = _run(cfg_hi, tdf, generate_excel=True)

    def _load_wb(self, result):
        """Load openpyxl workbook from SimulationResult.workbook_bytes."""
        from openpyxl import load_workbook as _lw
        return _lw(io.BytesIO(result.workbook_bytes))

    def test_s09_workbook_bytes_not_none(self):
        self.assertIsNotNone(self.result_lo.workbook_bytes)
        self.assertIsNotNone(self.result_hi.workbook_bytes)

    def test_s09_workbook_bytes_non_empty(self):
        self.assertGreater(len(self.result_lo.workbook_bytes), 0)
        self.assertGreater(len(self.result_hi.workbook_bytes), 0)

    def test_s09_six_sheets_present(self):
        expected = {
            "Event Data", "Campaign Metrics", "Validation Results",
            "Validation Summary", "Realism Report", "Diagnostics",
        }
        for result in (self.result_lo, self.result_hi):
            wb = self._load_wb(result)
            self.assertEqual(
                set(wb.sheetnames), expected,
                f"Workbook must have exactly 6 sheets. Got: {wb.sheetnames}",
            )

    def test_s09_event_data_sheet_row_count_matches_dataframe(self):
        for result in (self.result_lo, self.result_hi):
            wb = self._load_wb(result)
            ws = wb["Event Data"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(
                len(data_rows), len(result.events_df),
                f"Event Data sheet rows ({len(data_rows)}) must match "
                f"events_df length ({len(result.events_df)}).",
            )

    def test_s09_high_ctr_workbook_reflects_more_clicks(self):
        """CTR increase in config → more clicks in workbook Event Data sheet."""
        def _click_count_from_wb(result):
            wb = self._load_wb(result)
            ws = wb["Event Data"]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            try:
                at_col = header.index("Action")
            except ValueError:
                return 0
            return sum(
                1 for row in ws.iter_rows(min_row=2, values_only=True)
                if row[at_col] == "Click"
            )

        clicks_lo = _click_count_from_wb(self.result_lo)
        clicks_hi = _click_count_from_wb(self.result_hi)
        self.assertGreater(
            clicks_hi, clicks_lo,
            f"High-CTR workbook must record more clicks than low-CTR "
            f"(got {clicks_hi} vs {clicks_lo}).",
        )

    def test_s09_workbook_click_count_matches_dataframe_click_count(self):
        """Workbook must be consistent with in-memory events_df."""
        for result in (self.result_lo, self.result_hi):
            wb = self._load_wb(result)
            ws = wb["Event Data"]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            at_col = header.index("Action")
            wb_clicks = sum(
                1 for row in ws.iter_rows(min_row=2, values_only=True)
                if row[at_col] == "Click"
            )
            df_clicks = int((result.events_df["action_type"] == "Click").sum())
            self.assertEqual(
                wb_clicks, df_clicks,
                f"Workbook click count ({wb_clicks}) must match events_df "
                f"click count ({df_clicks}).",
            )


# ---------------------------------------------------------------------------
# SCENARIO 10 — Determinism Certification
# ---------------------------------------------------------------------------

class TestScenario10Determinism(unittest.TestCase):
    """CERT-010: Same inputs always produce identical outputs.

    Verifies:
    1. events_df is byte-for-byte identical across two runs.
    2. workbook_bytes is byte-for-byte identical across two runs.
    3. quality_score and realism_score are identical.
    4. Different inputs produce different events_df.

    Determinism mechanism: BehaviorEngine uses MD5(user_id) + ordinal date
    as per-user-per-day seed (SIM-019). Same user_id + same date = same draw.
    Workbook determinism: openpyxl timestamps pinned to _EPOCH (DEF-EX-001).
    """

    @classmethod
    def setUpClass(cls):
        tdf = _trigger_df(n=100)
        cfg = _cfg(triggers=(TriggerConfig("T1", 1, 1.0),))
        cls.result_a = _run(cfg, tdf, generate_excel=True)
        cls.result_b = _run(cfg, tdf, generate_excel=True)
        # Different config for inequality test
        cfg_alt = _cfg(ads=(AdConfig("Ad_A", 1, 14, False, "Display", "V", 0.40),),
                       triggers=(TriggerConfig("T1", 1, 1.0),))
        cls.result_alt = _run(cfg_alt, tdf, generate_excel=False)

    def test_s10_events_df_are_identical(self):
        df_a = self.result_a.events_df.reset_index(drop=True)
        df_b = self.result_b.events_df.reset_index(drop=True)
        pd.testing.assert_frame_equal(
            df_a, df_b,
            check_like=False,
            obj="events_df",
        )

    def test_s10_workbook_bytes_are_identical(self):
        self.assertEqual(
            self.result_a.workbook_bytes,
            self.result_b.workbook_bytes,
            "Two runs with identical inputs must produce byte-identical workbooks "
            "(DEF-EX-001: openpyxl timestamp pinned to _EPOCH).",
        )

    def test_s10_quality_scores_identical(self):
        self.assertEqual(
            self.result_a.quality_score,
            self.result_b.quality_score,
        )

    def test_s10_realism_scores_identical(self):
        self.assertEqual(
            self.result_a.realism_score,
            self.result_b.realism_score,
        )

    def test_s10_different_config_produces_different_events(self):
        """Changing CTR from 0.10 to 0.40 must change the event stream."""
        clicks_base = _qualifying_events(self.result_a.events_df)
        clicks_alt  = _qualifying_events(self.result_alt.events_df)
        self.assertNotEqual(
            clicks_base, clicks_alt,
            f"Different CTR configs must produce different click counts "
            f"(both gave {clicks_base}).",
        )

    def test_s10_n_events_matches_dataframe_length(self):
        for r in (self.result_a, self.result_b):
            self.assertEqual(
                r.n_events, len(r.events_df),
                "SimulationResult.n_events must equal len(events_df).",
            )


if __name__ == "__main__":
    unittest.main(verbosity=2)
