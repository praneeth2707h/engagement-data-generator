"""Stage 13 — Multi-Trigger End-to-End Certification Suite.

Proves that users belonging to multiple triggers are handled correctly
through every stage of the simulation pipeline.

Certification question (MT-CERT-Q-001):
    "Can users belong to multiple triggers without creating attribution
     errors, reporting errors, or journey assignment errors?"

Architecture under test
-----------------------
* ARCH-013 — Alphabetical Trigger_Name tiebreak for equal-priority triggers.
* ARCH-014 — Segment follows winning trigger row.
* ARCH-018 — Dropped journey → EXCLUDED regardless of trigger count.
* ARCH-020 — allow_reentry controls RE_ENTRY vs EXCLUDED for cooling users.

Priority resolution rule
------------------------
Lower `priority` integer = higher priority.
  TriggerConfig("T1", priority=1) beats TriggerConfig("T2", priority=2).
Tiebreak (equal priority): alphabetically first Trigger_Name wins.
  "T_Alpha" beats "T_Beta" at the same priority level.

Multi-trigger invariants certified here
----------------------------------------
1. Each user receives EXACTLY ONE winning trigger_name in state_df.
2. All events for a user are attributed to their single winning trigger.
3. TER is reported against the winning-trigger denominator, not the full file.
4. Re-entry priority resolution is identical to first-entry resolution.
5. Journey stage progression uses winning-trigger ads only.
6. ValidationEngine VAL-010 passes (no user in events with >1 trigger name).
7. Workbook Trigger_Name column reflects winning trigger for every row.
8. Identical multi-trigger inputs produce identical outputs (determinism).

Design notes
------------
* engagement_cooldown_days=0 used in most scenarios to maximise
  observable click events without confounding cooldown effects.
* All simulations use generate_excel=False except MT-009 (workbook) and
  MT-010 (determinism bytes check).
* N values are kept small (20–100) for fast CI execution (~25 s total).

References
----------
MT-001 .. MT-010  — Certification scenario IDs
ARCH-013, ARCH-014, ARCH-018, ARCH-020
VAL-009, VAL-010, VAL-003 (TER), VAL-013 (TCC)
DEF-MT-001 .. (populated if defects found)
"""
from __future__ import annotations

import io
import sys
import unittest
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from core.simulation_orchestrator import SimulationOrchestrator
from core.audience_manager import AudienceManager
from core.user_state_manager import UserStateManager
from core.validation_engine import ValidationEngine
from models.ad_config import AdConfig
from models.config_registry import ConfigRegistry
from models.enums import EligibilityStatus, JourneyStatus
from models.trigger_config import TriggerConfig
from tests.test_core.conftest import make_config


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_SIM_START = date(2024, 1, 1)   # Monday
_SIM_END_14 = date(2024, 1, 14)
_SIM_END_7  = date(2024, 1, 7)
_CAMPAIGN   = "TEST_CAMPAIGN"


def _base_cfg(**kw) -> ConfigRegistry:
    """Minimal two-trigger config with permissive caps."""
    defaults = dict(
        simulation_start_date=_SIM_START,
        simulation_end_date=_SIM_END_14,
        weekly_impression_cap=20,
        weekly_engagement_cap=14,
        weekly_click_cap=20,
        weekly_open_cap=20,
        engagement_cooldown_days=0,
        cooling_period_days=0,
        ads=(AdConfig("Ad_A", 1, 14, False, "Display", "VendorX", 0.10),),
        triggers=(
            TriggerConfig("T1", 1, 1.0),
            TriggerConfig("T2", 2, 1.0),
        ),
        segments=(),
        channels=(),
    )
    defaults.update(kw)
    return make_config(**defaults)


def _tdf_rows(*rows: tuple) -> pd.DataFrame:
    """Build trigger DataFrame from (User_ID, Trigger_Name, Segment) tuples."""
    return pd.DataFrame(
        [
            {
                "Campaign_ID":  _CAMPAIGN,
                "User_ID":      r[0],
                "Trigger_Name": r[1],
                "Segment":      r[2],
                "Trigger_Date": _SIM_START,
            }
            for r in rows
        ]
    )


def _run(
    cfg: ConfigRegistry,
    tdf: pd.DataFrame,
    previous_state_df: pd.DataFrame | None = None,
    generate_excel: bool = False,
):
    return SimulationOrchestrator(cfg).run(
        trigger_df=tdf,
        historical_df=None,
        previous_state_df=previous_state_df,
        generate_excel=generate_excel,
    )


def _winner_trigger(result, user_id: str) -> str:
    """Return the winning trigger_name for a specific user from audience_df."""
    row = result.audience_df[result.audience_df["user_id"] == user_id]
    return str(row["trigger_name"].iloc[0]) if len(row) else ""


def _event_triggers_for_user(result, user_id: str) -> set[str]:
    """Return the set of distinct trigger_names in events for a given user."""
    ev = result.events_df
    user_events = ev[ev["user_id"] == user_id]
    return set(user_events["trigger_name"].astype(str).unique())


# ---------------------------------------------------------------------------
# MT-001 — Single User, Single Trigger
# ---------------------------------------------------------------------------

class TestMT001SingleTrigger(unittest.TestCase):
    """MT-001: Baseline — single user assigned to T1 only.

    Establishes correct single-trigger behaviour before testing multi-trigger
    scenarios.  All subsequent scenarios must exceed or match these guarantees.
    """

    @classmethod
    def setUpClass(cls):
        tdf = _tdf_rows(
            ("U001", "T1", "Seg_A"),
            ("U002", "T1", "Seg_A"),
            ("U003", "T1", "Seg_A"),
        )
        cls.cfg = _base_cfg()
        cls.result = _run(cls.cfg, tdf)
        cls.audience = cls.result.audience_df
        cls.events   = cls.result.events_df

    def test_mt001_simulation_succeeds(self):
        self.assertTrue(self.result.succeeded, "MT-001: simulation must succeed.")

    def test_mt001_all_users_assigned_T1(self):
        triggers = self.audience["trigger_name"].astype(str).unique().tolist()
        self.assertEqual(
            triggers, ["T1"],
            f"MT-001: all users should be assigned T1; got {triggers}",
        )

    def test_mt001_events_attributed_to_T1_only(self):
        ev_triggers = self.events["trigger_name"].astype(str).unique().tolist()
        self.assertIn("T1", ev_triggers)
        self.assertNotIn(
            "T2", ev_triggers,
            "MT-001: no T2 events should exist when only T1 is in trigger file.",
        )

    def test_mt001_trigger_history_populated(self):
        for _, row in self.audience.iterrows():
            hist = str(row.get("trigger_history", ""))
            self.assertIn(
                "T1", hist,
                f"MT-001: trigger_history for {row['user_id']} must contain T1.",
            )

    def test_mt001_events_generated(self):
        self.assertGreater(len(self.events), 0, "MT-001: must produce events.")


# ---------------------------------------------------------------------------
# MT-002 — Single User, Two Triggers (Priority Resolution)
# ---------------------------------------------------------------------------

class TestMT002TwoTriggerPriorityResolution(unittest.TestCase):
    """MT-002: User in T1(priority=1) + T2(priority=2) — T1 must win.

    Priority rule: lower integer = higher priority.  T1 (priority=1) beats
    T2 (priority=2) for any user appearing in both rows.
    """

    @classmethod
    def setUpClass(cls):
        # U001: in both T1 and T2 → T1 wins
        # U002: T1 only
        # U003: T2 only
        tdf = _tdf_rows(
            ("U001", "T1", "Seg_A"),
            ("U001", "T2", "Seg_A"),
            ("U002", "T1", "Seg_A"),
            ("U003", "T2", "Seg_B"),
        )
        cls.cfg    = _base_cfg()
        cls.result = _run(cls.cfg, tdf)
        cls.audience = cls.result.audience_df
        cls.events   = cls.result.events_df

    def test_mt002_simulation_succeeds(self):
        self.assertTrue(self.result.succeeded)

    def test_mt002_shared_user_wins_T1(self):
        winner = _winner_trigger(self.result, "U001")
        self.assertEqual(
            winner, "T1",
            f"MT-002: U001 (in T1+T2) must win T1 (priority=1). Got {winner!r}.",
        )

    def test_mt002_shared_user_events_attributed_to_T1_only(self):
        ev_trigs = _event_triggers_for_user(self.result, "U001")
        self.assertEqual(
            ev_trigs, {"T1"},
            f"MT-002: U001 events must be T1-only; got {ev_trigs}.",
        )

    def test_mt002_T1_only_user_stays_T1(self):
        self.assertEqual(_winner_trigger(self.result, "U002"), "T1")

    def test_mt002_T2_only_user_stays_T2(self):
        self.assertEqual(_winner_trigger(self.result, "U003"), "T2")

    def test_mt002_each_user_has_exactly_one_trigger_in_events(self):
        ev = self.events
        for uid in ev["user_id"].unique():
            trigs = set(ev.loc[ev["user_id"] == uid, "trigger_name"].astype(str))
            self.assertEqual(
                len(trigs), 1,
                f"MT-002: user {uid} must have exactly 1 trigger in events; got {trigs}.",
            )

    def test_mt002_total_user_count_correct(self):
        # 3 distinct users
        self.assertEqual(self.result.n_users, 3)


# ---------------------------------------------------------------------------
# MT-003 — Three Triggers, Highest Priority Wins + Alphabetical Tiebreak
# ---------------------------------------------------------------------------

class TestMT003ThreeTriggerPriorityAndTiebreak(unittest.TestCase):
    """MT-003: User in T1(priority=3), T2(priority=1), T3(priority=2) — T2 wins.

    Also certifies the alphabetical tiebreak (ARCH-013): when two triggers
    share the same priority, the alphabetically first Trigger_Name wins.
    """

    @classmethod
    def setUpClass(cls):
        # Three-way: U001 in all three; T2 wins (priority=1)
        cfg3 = _base_cfg(
            triggers=(
                TriggerConfig("T1", 3, 1.0),
                TriggerConfig("T2", 1, 1.0),
                TriggerConfig("T3", 2, 1.0),
            )
        )
        tdf3 = _tdf_rows(
            ("U001", "T1", "Seg_A"),
            ("U001", "T2", "Seg_A"),
            ("U001", "T3", "Seg_A"),
            ("U002", "T3", "Seg_A"),  # T3 only
        )
        cls.result3 = _run(cfg3, tdf3)

        # Tiebreak: U001 in T_Alpha(priority=1) + T_Beta(priority=1) → T_Alpha wins
        cfg_tie = _base_cfg(
            triggers=(
                TriggerConfig("T_Alpha", 1, 1.0),
                TriggerConfig("T_Beta",  1, 1.0),
            )
        )
        tdf_tie = _tdf_rows(
            ("U001", "T_Beta",  "Seg_A"),
            ("U001", "T_Alpha", "Seg_A"),
        )
        cls.result_tie = _run(cfg_tie, tdf_tie)

    def test_mt003_simulation_succeeds(self):
        self.assertTrue(self.result3.succeeded)
        self.assertTrue(self.result_tie.succeeded)

    def test_mt003_T2_wins_lowest_priority_number(self):
        winner = _winner_trigger(self.result3, "U001")
        self.assertEqual(
            winner, "T2",
            f"MT-003: T2(priority=1) must beat T1(3) and T3(2). Got {winner!r}.",
        )

    def test_mt003_U002_stays_T3(self):
        winner = _winner_trigger(self.result3, "U002")
        self.assertEqual(winner, "T3")

    def test_mt003_events_show_only_winning_trigger(self):
        ev_trigs = _event_triggers_for_user(self.result3, "U001")
        self.assertEqual(
            ev_trigs, {"T2"},
            f"MT-003: U001 events must be T2 only; got {ev_trigs}.",
        )

    def test_mt003_alphabetical_tiebreak_T_Alpha_wins(self):
        winner = _winner_trigger(self.result_tie, "U001")
        self.assertEqual(
            winner, "T_Alpha",
            f"MT-003: T_Alpha must beat T_Beta on alphabetical tiebreak. Got {winner!r}.",
        )

    def test_mt003_tiebreak_events_T_Alpha_only(self):
        ev_trigs = _event_triggers_for_user(self.result_tie, "U001")
        self.assertEqual(ev_trigs, {"T_Alpha"})


# ---------------------------------------------------------------------------
# MT-004 — Mixed Population Attribution Counts
# ---------------------------------------------------------------------------

class TestMT004MixedPopulationAttribution(unittest.TestCase):
    """MT-004: Mixed population — T1-only, T2-only, and both-trigger users.

    Population:
        40 users in T1 only  → all win T1
        40 users in T2 only  → all win T2
        20 users in T1 + T2  → all win T1 (T1 has priority=1)

    Expected attribution:
        T1: 40 + 20 = 60 users
        T2: 40       = 40 users
        Total: 100 unique users, zero in both triggers simultaneously.
    """

    _N_T1   = 40
    _N_T2   = 40
    _N_BOTH = 20

    @classmethod
    def setUpClass(cls):
        rows = []
        for i in range(cls._N_T1):
            rows.append((f"U_T1_{i:03d}", "T1", "Seg_A"))
        for i in range(cls._N_T2):
            rows.append((f"U_T2_{i:03d}", "T2", "Seg_B"))
        for i in range(cls._N_BOTH):
            rows.append((f"U_MT_{i:03d}", "T1", "Seg_A"))
            rows.append((f"U_MT_{i:03d}", "T2", "Seg_B"))

        tdf = _tdf_rows(*rows)
        cls.cfg    = _base_cfg()
        cls.result = _run(cls.cfg, tdf)
        cls.aud    = cls.result.audience_df
        cls.events = cls.result.events_df

    def test_mt004_simulation_succeeds(self):
        self.assertTrue(self.result.succeeded)

    def test_mt004_T1_gets_60_users(self):
        t1_users = (self.aud["trigger_name"].astype(str) == "T1").sum()
        self.assertEqual(
            t1_users, self._N_T1 + self._N_BOTH,
            f"MT-004: T1 should have {self._N_T1 + self._N_BOTH} users; got {t1_users}.",
        )

    def test_mt004_T2_gets_40_users(self):
        t2_users = (self.aud["trigger_name"].astype(str) == "T2").sum()
        self.assertEqual(
            t2_users, self._N_T2,
            f"MT-004: T2 should have {self._N_T2} users; got {t2_users}.",
        )

    def test_mt004_total_user_count_100(self):
        self.assertEqual(self.result.n_users, self._N_T1 + self._N_T2 + self._N_BOTH)

    def test_mt004_no_user_attributed_to_both_triggers(self):
        """Partition must be complete — no user appears in both T1 and T2 groups."""
        t1_set = set(self.aud.loc[self.aud["trigger_name"].astype(str) == "T1", "user_id"])
        t2_set = set(self.aud.loc[self.aud["trigger_name"].astype(str) == "T2", "user_id"])
        overlap = t1_set & t2_set
        self.assertEqual(
            len(overlap), 0,
            f"MT-004: {len(overlap)} user(s) appear in both T1 and T2 groups: {overlap}.",
        )

    def test_mt004_both_trigger_users_win_T1(self):
        """All 20 users who appear in both triggers must be attributed to T1."""
        for i in range(self._N_BOTH):
            uid = f"U_MT_{i:03d}"
            winner = _winner_trigger(self.result, uid)
            self.assertEqual(
                winner, "T1",
                f"MT-004: multi-trigger user {uid} should win T1; got {winner!r}.",
            )

    def test_mt004_val010_passes(self):
        """VAL-010: no user appears with >1 trigger_name in events."""
        vr, _, _ = ValidationEngine(self.cfg).validate(self.events, self.aud)
        mt_row = vr[vr["rule_name"].astype(str).str.contains("Multi-Trigger", na=False)]
        self.assertFalse(mt_row.empty, "MT-004: VAL-010 Multi-Trigger row must exist.")
        self.assertEqual(
            str(mt_row["status"].iloc[0]), "Pass",
            f"MT-004: VAL-010 must Pass; got {mt_row['status'].iloc[0]}.",
        )


# ---------------------------------------------------------------------------
# MT-005 — TER Reporting Against Correct Trigger
# ---------------------------------------------------------------------------

class TestMT005TERReporting(unittest.TestCase):
    """MT-005: TER is computed and reported against the correct trigger.

    Verifies that:
    - ValidationEngine produces one TER rule row per trigger.
    - Each TER row uses the users who WON that trigger as the denominator.
    - T2's TER denominator = 40 (T2-only users), NOT 60 (T2-only + both-trigger).
    """

    _N_T1   = 40
    _N_T2   = 40
    _N_BOTH = 20

    @classmethod
    def setUpClass(cls):
        rows = []
        for i in range(cls._N_T1):
            rows.append((f"U_T1_{i:03d}", "T1", "Seg_A"))
        for i in range(cls._N_T2):
            rows.append((f"U_T2_{i:03d}", "T2", "Seg_B"))
        for i in range(cls._N_BOTH):
            rows.append((f"U_MT_{i:03d}", "T1", "Seg_A"))
            rows.append((f"U_MT_{i:03d}", "T2", "Seg_B"))

        tdf = _tdf_rows(*rows)
        cls.cfg    = _base_cfg()
        cls.result = _run(cls.cfg, tdf)
        cls.vr, cls.vs, cls.rr = ValidationEngine(cls.cfg).validate(
            cls.result.events_df, cls.result.audience_df
        )

    def test_mt005_simulation_succeeds(self):
        self.assertTrue(self.result.succeeded)

    def test_mt005_ter_rows_exist_for_T1_and_T2(self):
        ter_names = set(
            self.vr.loc[self.vr["rule_name"].astype(str).str.startswith("TER"), "rule_name"]
            .astype(str)
        )
        self.assertIn("TER Achievement — T1", ter_names)
        self.assertIn("TER Achievement — T2", ter_names)

    def test_mt005_T1_ter_denominator_is_60(self):
        """T1's denominator in TER message = 60 (T1-only 40 + both-trigger 20)."""
        t1_row = self.vr[self.vr["rule_name"].astype(str) == "TER Achievement — T1"]
        self.assertFalse(t1_row.empty)
        msg = str(t1_row["message"].iloc[0])
        # Message contains "XX/60 users" or similar
        self.assertIn(
            "/60 users", msg,
            f"MT-005: T1 TER message must reference 60 users; got: {msg}",
        )

    def test_mt005_T2_ter_denominator_is_40(self):
        """T2's denominator = 40 (T2-only users); multi-trigger users went to T1."""
        t2_row = self.vr[self.vr["rule_name"].astype(str) == "TER Achievement — T2"]
        self.assertFalse(t2_row.empty)
        msg = str(t2_row["message"].iloc[0])
        self.assertIn(
            "/40 users", msg,
            f"MT-005: T2 TER message must reference 40 users; got: {msg}",
        )

    def test_mt005_val010_passes(self):
        mt_row = self.vr[self.vr["rule_name"].astype(str).str.contains("Multi-Trigger", na=False)]
        self.assertFalse(mt_row.empty)
        self.assertEqual(str(mt_row["status"].iloc[0]), "Pass")


# ---------------------------------------------------------------------------
# MT-006 — Re-entry User with Multiple Triggers
# ---------------------------------------------------------------------------

class TestMT006ReEntryMultiTrigger(unittest.TestCase):
    """MT-006: Re-entry users appearing in multiple triggers still get priority resolution.

    Setup:
        20 users pre-loaded with expired cooling (journey_status=Completed).
        Each of the 20 appears in BOTH T1 and T2 rows in the new trigger file.
        allow_reentry=True → they become RE_ENTRY; T1 still wins (priority=1).

    Verifies:
        - Re-entry users are correctly classified RE_ENTRY.
        - Their winning trigger is T1 (not T2).
        - Disabling re-entry excludes them correctly.
    """

    _N_NEW    = 30  # New users (T1 only)
    _N_REENTRY = 20  # Cooling-expired users, in both T1 and T2

    @classmethod
    def setUpClass(cls):
        new_users     = [f"U_NEW_{i:03d}" for i in range(cls._N_NEW)]
        cooling_users = [f"U_COOL_{i:03d}" for i in range(cls._N_REENTRY)]

        rows = []
        for u in new_users:
            rows.append((u, "T1", "Seg_A"))
        for u in cooling_users:
            rows.append((u, "T1", "Seg_A"))  # T1 and T2 both for re-entry users
            rows.append((u, "T2", "Seg_B"))

        tdf = _tdf_rows(*rows)
        cls.tdf = tdf

        # Build previous state with cooling users
        cfg_init = _base_cfg(cooling_period_days=30, allow_reentry=True)
        usm = UserStateManager(cfg_init)
        prev = usm.initialize_user_states(tdf, previous_state_df=None).copy()

        # Set cooling users to expired cooling with completed journey
        cool_idx = prev[prev["user_id"].isin(cooling_users)].index
        prev.loc[cool_idx, "eligibility_status"] = EligibilityStatus.COOLING.value
        prev.loc[cool_idx, "journey_status"]     = JourneyStatus.COMPLETED.value
        prev.loc[cool_idx, "cooling_period_end"] = date(2023, 12, 1)  # expired

        cls.prev_state = prev
        cls.cooling_users = cooling_users
        cls.new_users = new_users

        cfg_on  = _base_cfg(cooling_period_days=30, allow_reentry=True)
        cfg_off = _base_cfg(cooling_period_days=30, allow_reentry=False)

        cls.result_on  = _run(cfg_on,  tdf, previous_state_df=prev.copy())
        cls.result_off = _run(cfg_off, tdf, previous_state_df=prev.copy())

    def test_mt006_both_simulations_succeed(self):
        self.assertTrue(self.result_on.succeeded)
        self.assertTrue(self.result_off.succeeded)

    def test_mt006_reentry_on_classifies_cooling_users_as_RE_ENTRY(self):
        re_entry_count = (
            self.result_on.audience_df["eligibility_status"].astype(str)
            == EligibilityStatus.RE_ENTRY.value
        ).sum()
        self.assertEqual(
            re_entry_count, self._N_REENTRY,
            f"MT-006: expected {self._N_REENTRY} RE_ENTRY users; got {re_entry_count}.",
        )

    def test_mt006_reentry_users_win_T1(self):
        """Re-entry users in both T1+T2 must still win T1 (priority=1)."""
        aud = self.result_on.audience_df
        re_entry_users = aud[
            aud["eligibility_status"].astype(str) == EligibilityStatus.RE_ENTRY.value
        ]
        triggers = re_entry_users["trigger_name"].astype(str).unique().tolist()
        self.assertEqual(
            triggers, ["T1"],
            f"MT-006: re-entry users must win T1; got {triggers}.",
        )

    def test_mt006_reentry_off_excludes_cooling_users(self):
        excluded_count = (
            self.result_off.audience_df["eligibility_status"].astype(str)
            == EligibilityStatus.EXCLUDED.value
        ).sum()
        self.assertEqual(
            excluded_count, self._N_REENTRY,
            f"MT-006: {self._N_REENTRY} cooling-expired users should be EXCLUDED.",
        )

    def test_mt006_reentry_on_generates_more_events_than_off(self):
        self.assertGreater(
            self.result_on.n_events, self.result_off.n_events,
            "MT-006: re-entry ON must produce more events than OFF.",
        )

    def test_mt006_reentry_events_attributed_to_T1(self):
        """Events for re-entry users must be attributed to their winning trigger T1."""
        ev = self.result_on.events_df
        for uid in self.cooling_users:
            trigs = _event_triggers_for_user(self.result_on, uid)
            if trigs:  # some re-entry users may not generate events
                self.assertEqual(
                    trigs, {"T1"},
                    f"MT-006: re-entry user {uid} events must be T1; got {trigs}.",
                )


# ---------------------------------------------------------------------------
# MT-007 — Journey Assignment to Winning Trigger
# ---------------------------------------------------------------------------

class TestMT007JourneyAssignmentToWinningTrigger(unittest.TestCase):
    """MT-007: Journey progression uses the winning trigger's ad sequence.

    Setup:
        - 3-ad journey: Ad_A (dur=3), Ad_B (dur=3), Ad_C (dur=3)
        - All users appear in T1 + T2 — T1 wins
        - 14-day simulation: users should progress through Ad_A → Ad_B → Ad_C

    Verifies:
        - Events are generated using the winning trigger (T1).
        - Journey stages Ad_B and Ad_C appear (proving full progression).
        - No events are attributed to T2.
    """

    @classmethod
    def setUpClass(cls):
        cfg = _base_cfg(
            ads=(
                AdConfig("Ad_A", 1, 3,  False, "Display", "V", 0.20),
                AdConfig("Ad_B", 2, 3,  False, "Display", "V", 0.20),
                AdConfig("Ad_C", 3, 14, False, "Display", "V", 0.20),
            ),
            triggers=(
                TriggerConfig("T1", 1, 1.0),
                TriggerConfig("T2", 2, 1.0),
            ),
        )
        rows = []
        for i in range(30):
            rows.append((f"U{i:03d}", "T1", "Seg_A"))
            rows.append((f"U{i:03d}", "T2", "Seg_A"))
        tdf = _tdf_rows(*rows)

        cls.cfg    = cfg
        cls.result = _run(cfg, tdf)
        cls.events = cls.result.events_df

    def test_mt007_simulation_succeeds(self):
        self.assertTrue(self.result.succeeded)

    def test_mt007_no_T2_events(self):
        t2_events = (self.events["trigger_name"].astype(str) == "T2").sum()
        self.assertEqual(
            t2_events, 0,
            f"MT-007: no events should be attributed to T2 when T1 wins; got {t2_events}.",
        )

    def test_mt007_T1_events_exist(self):
        t1_events = (self.events["trigger_name"].astype(str) == "T1").sum()
        self.assertGreater(t1_events, 0, "MT-007: T1 events must exist.")

    def test_mt007_journey_advances_to_Ad_B(self):
        """3-day Ad_A → users should reach Ad_B during 14-day simulation."""
        ad_b_events = (self.events["current_ad"].astype(str) == "Ad_B").sum()
        self.assertGreater(
            ad_b_events, 0,
            f"MT-007: users should advance to Ad_B; got {ad_b_events} Ad_B events.",
        )

    def test_mt007_journey_advances_to_Ad_C(self):
        """After 6 days (3+3), users should reach Ad_C."""
        ad_c_events = (self.events["current_ad"].astype(str) == "Ad_C").sum()
        self.assertGreater(
            ad_c_events, 0,
            f"MT-007: users should advance to Ad_C; got {ad_c_events} Ad_C events.",
        )

    def test_mt007_all_events_use_winning_trigger(self):
        """Every event must carry the winning trigger T1, never T2."""
        unique_triggers = set(self.events["trigger_name"].astype(str).unique())
        self.assertEqual(
            unique_triggers, {"T1"},
            f"MT-007: all events must use T1; found {unique_triggers}.",
        )


# ---------------------------------------------------------------------------
# MT-008 — ValidationEngine Multi-Trigger Correctness
# ---------------------------------------------------------------------------

class TestMT008ValidationEngineMultiTrigger(unittest.TestCase):
    """MT-008: ValidationEngine rules remain correct in multi-trigger scenarios.

    Checks:
        VAL-009 — No unknown trigger names in events or state.
        VAL-010 — Each user has exactly one trigger_name in events.
        VAL-003 — TER Achievement rows exist per trigger.
        VAL-013 — TCC Calculation rows exist per trigger.
    """

    _N_T1   = 30
    _N_T2   = 30
    _N_BOTH = 20

    @classmethod
    def setUpClass(cls):
        rows = []
        for i in range(cls._N_T1):
            rows.append((f"U_T1_{i:03d}", "T1", "Seg_A"))
        for i in range(cls._N_T2):
            rows.append((f"U_T2_{i:03d}", "T2", "Seg_B"))
        for i in range(cls._N_BOTH):
            rows.append((f"U_MT_{i:03d}", "T1", "Seg_A"))
            rows.append((f"U_MT_{i:03d}", "T2", "Seg_B"))

        tdf = _tdf_rows(*rows)
        cls.cfg    = _base_cfg()
        cls.result = _run(cls.cfg, tdf)
        cls.vr, cls.vs, cls.rr = ValidationEngine(cls.cfg).validate(
            cls.result.events_df, cls.result.audience_df
        )
        cls.rule_names = set(cls.vr["rule_name"].astype(str))

    def test_mt008_simulation_succeeds(self):
        self.assertTrue(self.result.succeeded)

    def test_mt008_val010_multi_trigger_passes(self):
        """VAL-010: no user appears with >1 trigger_name in events."""
        mt_row = self.vr[self.vr["rule_name"].astype(str).str.contains("Multi-Trigger", na=False)]
        self.assertFalse(mt_row.empty, "MT-008: VAL-010 Multi-Trigger row must exist.")
        self.assertEqual(
            str(mt_row["status"].iloc[0]), "Pass",
            f"MT-008: VAL-010 must Pass; got {mt_row['status'].iloc[0]}.",
        )

    def test_mt008_val009_trigger_priority_passes(self):
        """VAL-009: all trigger names in events/state are known triggers."""
        tp_row = self.vr[self.vr["rule_name"].astype(str).str.contains("Trigger Priority", na=False)]
        self.assertFalse(tp_row.empty, "MT-008: VAL-009 Trigger Priority row must exist.")
        status = str(tp_row["status"].iloc[0])
        self.assertEqual(
            status, "Pass",
            f"MT-008: VAL-009 must Pass with known triggers; got {status}.",
        )

    def test_mt008_ter_rows_exist_for_T1_and_T2(self):
        self.assertIn("TER Achievement — T1", self.rule_names)
        self.assertIn("TER Achievement — T2", self.rule_names)

    def test_mt008_tcc_rows_exist_for_T1_and_T2(self):
        self.assertIn("TCC Calculation — T1", self.rule_names)
        self.assertIn("TCC Calculation — T2", self.rule_names)

    def test_mt008_quality_score_in_range(self):
        self.assertGreaterEqual(self.result.quality_score, 0.0)
        self.assertLessEqual(self.result.quality_score, 100.0)


# ---------------------------------------------------------------------------
# MT-009 — Workbook Export Trigger Attribution
# ---------------------------------------------------------------------------

class TestMT009WorkbookTriggerAttribution(unittest.TestCase):
    """MT-009: Downloaded workbook reflects winning trigger attribution correctly.

    Verifies:
        - Event Data sheet Trigger_Name column is populated.
        - T1 and T2 both appear in the workbook (for users with different winners).
        - No user appears under multiple triggers in the workbook.
        - Event row count in workbook matches events_df.
    """

    @classmethod
    def setUpClass(cls):
        rows = []
        for i in range(20):
            rows.append((f"U_T1_{i:03d}", "T1", "Seg_A"))  # T1 only
        for i in range(20):
            rows.append((f"U_T2_{i:03d}", "T2", "Seg_B"))  # T2 only
        for i in range(10):
            rows.append((f"U_MT_{i:03d}", "T1", "Seg_A"))  # T1 wins
            rows.append((f"U_MT_{i:03d}", "T2", "Seg_B"))

        tdf = _tdf_rows(*rows)
        cls.cfg    = _base_cfg()
        cls.result = _run(cls.cfg, tdf, generate_excel=True)

    def _load_wb(self):
        from openpyxl import load_workbook as _lw
        return _lw(io.BytesIO(self.result.workbook_bytes))

    def test_mt009_workbook_bytes_not_none(self):
        self.assertIsNotNone(self.result.workbook_bytes)
        self.assertGreater(len(self.result.workbook_bytes), 0)

    def test_mt009_event_data_row_count_matches_dataframe(self):
        wb = self._load_wb()
        ws = wb["Event Data"]
        data_rows = ws.max_row - 1  # minus header
        self.assertEqual(data_rows, len(self.result.events_df))

    def test_mt009_trigger_name_column_exists_in_workbook(self):
        wb = self._load_wb()
        ws = wb["Event Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn(
            "Trigger_Name", header,
            f"MT-009: Trigger_Name column must be in Event Data header; got {header}.",
        )

    def test_mt009_T1_and_T2_both_appear_in_workbook(self):
        wb = self._load_wb()
        ws = wb["Event Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        trig_col = header.index("Trigger_Name")
        triggers_in_wb = {
            row[trig_col]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[trig_col] is not None
        }
        self.assertIn("T1", triggers_in_wb)
        self.assertIn("T2", triggers_in_wb)

    def test_mt009_no_user_has_multiple_triggers_in_workbook(self):
        """Each user_id in the workbook must map to exactly one Trigger_Name."""
        wb = self._load_wb()
        ws = wb["Event Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        uid_col  = header.index("User_ID")
        trig_col = header.index("Trigger_Name")

        user_triggers: dict[str, set] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            uid  = row[uid_col]
            trig = row[trig_col]
            if uid and trig:
                user_triggers.setdefault(uid, set()).add(trig)

        violations = {u: t for u, t in user_triggers.items() if len(t) > 1}
        self.assertEqual(
            len(violations), 0,
            f"MT-009: {len(violations)} user(s) have >1 trigger in workbook: {violations}.",
        )

    def test_mt009_workbook_T1_events_reflect_T1_users_count(self):
        """T1 events in workbook should cover T1-only (20) + both-trigger (10) users."""
        wb = self._load_wb()
        ws = wb["Event Data"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        trig_col = header.index("Trigger_Name")
        uid_col  = header.index("User_ID")
        t1_users = {
            row[uid_col]
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[trig_col] == "T1"
        }
        self.assertEqual(
            len(t1_users), 30,  # 20 T1-only + 10 both-trigger users who won T1
            f"MT-009: T1 should have 30 distinct users in workbook; got {len(t1_users)}.",
        )


# ---------------------------------------------------------------------------
# MT-010 — Determinism with Multi-Trigger Inputs
# ---------------------------------------------------------------------------

class TestMT010Determinism(unittest.TestCase):
    """MT-010: Identical multi-trigger inputs always produce identical outputs.

    Also verifies that changing trigger priority order produces different
    attribution — confirming priority resolution is the mechanism, not luck.
    """

    @classmethod
    def setUpClass(cls):
        rows = []
        for i in range(20):
            rows.append((f"U_T1_{i:03d}", "T1", "Seg_A"))
        for i in range(20):
            rows.append((f"U_T2_{i:03d}", "T2", "Seg_B"))
        for i in range(10):
            rows.append((f"U_MT_{i:03d}", "T1", "Seg_A"))
            rows.append((f"U_MT_{i:03d}", "T2", "Seg_B"))

        cls.tdf = _tdf_rows(*rows)
        cls.cfg = _base_cfg()

        cls.result_a = _run(cls.cfg, cls.tdf, generate_excel=True)
        cls.result_b = _run(cls.cfg, cls.tdf, generate_excel=True)

        # Reversed-priority config: T2 now has priority=1, T1 has priority=2
        cfg_rev = _base_cfg(
            triggers=(
                TriggerConfig("T1", 2, 1.0),
                TriggerConfig("T2", 1, 1.0),
            )
        )
        cls.result_rev = _run(cfg_rev, cls.tdf, generate_excel=False)

    def test_mt010_events_df_identical(self):
        pd.testing.assert_frame_equal(
            self.result_a.events_df.reset_index(drop=True),
            self.result_b.events_df.reset_index(drop=True),
            obj="events_df",
        )

    def test_mt010_workbook_bytes_identical(self):
        self.assertEqual(
            self.result_a.workbook_bytes,
            self.result_b.workbook_bytes,
            "MT-010: identical multi-trigger inputs must produce byte-identical workbooks.",
        )

    def test_mt010_trigger_attribution_identical(self):
        aud_a = self.result_a.audience_df[["user_id", "trigger_name"]].sort_values("user_id")
        aud_b = self.result_b.audience_df[["user_id", "trigger_name"]].sort_values("user_id")
        pd.testing.assert_frame_equal(
            aud_a.reset_index(drop=True),
            aud_b.reset_index(drop=True),
            check_like=True,
            obj="audience_df trigger attribution",
        )

    def test_mt010_quality_scores_identical(self):
        self.assertEqual(self.result_a.quality_score, self.result_b.quality_score)

    def test_mt010_reversed_priority_changes_attribution(self):
        """Reversing T1↔T2 priority must change which trigger wins for multi-trigger users."""
        aud_orig = self.result_a.audience_df
        aud_rev  = self.result_rev.audience_df

        # Multi-trigger users: U_MT_000..009
        mt_users = [f"U_MT_{i:03d}" for i in range(10)]
        orig_winners = {
            str(aud_orig.loc[aud_orig["user_id"] == u, "trigger_name"].iloc[0])
            for u in mt_users
        }
        rev_winners  = {
            str(aud_rev.loc[aud_rev["user_id"] == u, "trigger_name"].iloc[0])
            for u in mt_users
        }
        self.assertEqual(orig_winners, {"T1"},
                         f"MT-010: original config must assign T1 to all MT users; got {orig_winners}")
        self.assertEqual(rev_winners,  {"T2"},
                         f"MT-010: reversed config must assign T2 to all MT users; got {rev_winners}")


if __name__ == "__main__":
    unittest.main(verbosity=2)
